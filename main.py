import sys, os, re
import polars as pl
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QLineEdit, QPushButton, QTableView, QHeaderView, QFileDialog, QRadioButton,
                               QComboBox, QMenu, QInputDialog, QMessageBox, QProgressBar, QGridLayout,
                               QScrollArea, QFrame, QDateEdit, QSizePolicy, QCheckBox, QAbstractItemView,
                               QCompleter, QDialog, QDialogButtonBox, QButtonGroup)
from PySide6.QtCore import Qt, QAbstractTableModel, QDate, QThread, Signal, Slot, QTimer, QUrl, QStringListModel, QSize, QItemSelectionModel
from PySide6.QtGui import QPainter, QAction, QIcon, QFont, QColor, QDesktopServices, QPixmap

import numpy as np # For .npz files
import pickle      # For .pkl files
import hickle
import time
from pathlib import Path
from chardet import detect
from csv import Sniffer, QUOTE_MINIMAL, excel
from datetime import datetime
import tempfile
from datetime import datetime, date
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import codecs
from pathlib import Path
from openpyxl import load_workbook
from tabulate import tabulate
import xlrd
from operation_logger import logger 
import threading

if "NUITKA_ONEFILE_PARENT" in os.environ:
   splash_filename = os.path.join(
      tempfile.gettempdir(),
      "onefile_%d_splash_feedback.tmp" % int(os.environ["NUITKA_ONEFILE_PARENT"]),
   )

   if os.path.exists(splash_filename):
      os.unlink(splash_filename)

def timer(func):
    def wrapper(*args, **kwargs):
        start = time.time()
        res = func(*args, **kwargs)
        print(f"\n >> {func.__name__} took {time.time() - start} seconds <<\n")
        return res
    return wrapper

pl.when(pl.col("POL_NB").cast(pl.Utf8).fill_null("").str.contains(r"^(?i)\d+$")).then(pl.lit('Manual (FC)')).otherwise(
    pl.when(pl.col("POL_NB").cast(pl.Utf8).fill_null("").str.contains(r"^(?i)\w{2}\d+\w{2}\d\w\d{2}$")).then(pl.lit('Portfolio (Country, FC, MPCode)')).otherwise(
        pl.when(pl.col("POL_NB").cast(pl.Utf8).fill_null("").str.contains(r"^(?i)DE\d+F6\w{2}\d\w\d{2}$")).then(pl.lit('Portfolio (Country, FC, F6, MPCode)')).otherwise(pl.lit('Unidentified'))
        )
)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller/Nuitka """
    if hasattr(sys, r'_MEIPASS') :
        try:
            # PyInstaller/Nuitka creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS # type: ignore
        except AttributeError: # AttributeError if _MEIPASS is not set (e.g. running as script)
            base_path = os.path.abspath(".")
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def get_sheet_names(file_name):
    if file_name.lower().endswith('.xls'):
        with xlrd.open_workbook(file_name, on_demand=True) as workbook:
            return workbook.sheet_names()
    elif file_name.lower().endswith('.xlsb'):
        try:
            from pyxlsb import open_workbook as open_xlsb_workbook
            with open_xlsb_workbook(file_name) as workbook:
                return workbook.sheets
        except ImportError:
            raise ValueError("pyxlsb is required for .xlsb files. Please install it.")
    elif file_name.lower().endswith(('.xlsx', '.xlsm')):
        try:
            with ZipFile(file_name) as archive:
                if 'xl/workbook.xml' not in archive.namelist():
                    return load_workbook(file_name, read_only=True, data_only=True).sheetnames
                tree = ET.parse(archive.open('xl/workbook.xml'))
                root = tree.getroot()
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                sheets_data = []
                for sheet_node in root.findall('.//main:sheet', ns):
                    name = sheet_node.get('name')
                    sheets_data.append({'name': name})
                return [s['name'] for s in sheets_data]
        except Exception:
            return load_workbook(file_name, read_only=True, data_only=True) .sheetnames
    else:
        raise ValueError(f"Unsupported file format: {file_name}")

class ContainerFileNavigator:
    """
    Helper class to inspect and select data from container-like file formats
    such as NPZ, Pickle, or Hickle files that might store multiple datasets
    or complex structures.
    """
    def __init__(self, parent_window=None):
        # parent_window is needed if we need to show QInputDialog from here,
        # though it's better for the worker to signal main thread for UI.
        # For now, this class will return information that worker thread can use.
        self.parent_window = parent_window 

    def inspect_npz(self, file_path):
        """Returns a list of array keys and their shapes/dtypes from an NPZ file."""
        try:
            with np.load(file_path) as data:
                keys = list(data.keys())
                if not keys:
                    return []
                content_info = []
                for key in keys:
                    array_preview = data[key]
                    content_info.append({
                        "name": key, 
                        "shape": array_preview.shape, 
                        "dtype": str(array_preview.dtype)
                    })
                return content_info
        except Exception as e:
            print(f"Error inspecting NPZ file {file_path}: {e}")
            return [{"name": f"Error: {e}", "shape": (), "dtype": "Error"}]

    def load_selected_npz_array(self, file_path, array_name):
        """Loads a specific array from an NPZ file into a Polars DataFrame."""
        try:
            with np.load(file_path) as data:
                if array_name not in data:
                    raise ValueError(f"Array '{array_name}' not found in NPZ file.")
                loaded_array = data[array_name]

                if loaded_array.ndim == 1:
                    return pl.DataFrame({array_name: loaded_array})
                elif loaded_array.ndim == 2:
                    col_names = [f"{array_name}_{i}" for i in range(loaded_array.shape[1])]
                    return pl.DataFrame(loaded_array, schema=col_names)
                elif loaded_array.ndim == 0 and loaded_array.dtype.fields: # Structured array
                    return pl.from_numpy(loaded_array, orient="row")
                else:
                    raise ValueError(f"Selected array '{array_name}' is not 1D/2D or recognized structured array.")
        except Exception as e:
            print(f"Error loading array '{array_name}' from NPZ {file_path}: {e}")
            raise # Re-raise for the worker to catch

    def _inspect_python_object(self, obj, current_path=""):
        """
        Recursively inspects a Python object (from Pickle/Hickle)
        to find Polars/Pandas DataFrames or convertible structures.
        Returns a list of dicts: {"name": "path.to.df", "type": "Polars DataFrame", "shape": (r,c)}
        """
        items = []
        if isinstance(obj, pl.DataFrame):
            items.append({"name": current_path or "DataFrame", "type": "Polars DataFrame", "shape": obj.shape, "object_path": current_path})
        elif 'pandas' in sys.modules and isinstance(obj, sys.modules['pandas'].DataFrame):
            items.append({"name": current_path or "DataFrame", "type": "Pandas DataFrame", "shape": obj.shape, "object_path": current_path})
        elif isinstance(obj, np.ndarray) and (obj.ndim <= 2 or (obj.ndim == 0 and obj.dtype.fields)):
             items.append({"name": current_path or "NumPy Array", "type": "NumPy Array", "shape": obj.shape, "object_path": current_path})
        elif isinstance(obj, dict):
            for key, value in obj.items():
                new_path = f"{current_path}.{key}" if current_path else key
                items.extend(self._inspect_python_object(value, new_path))
        elif isinstance(obj, list) and obj: # Only inspect non-empty lists
            # If list of dicts (potential for DataFrame) or list of DataFrames
            if all(isinstance(item, dict) for item in obj) or \
               all((isinstance(item, (pl.DataFrame, sys.modules['pandas'].DataFrame if 'pandas' in sys.modules else object)) for item in obj)):
                 # For now, mark the list itself. User would then pick this "list of..."
                 # A more advanced loader might try to combine list of dicts/dfs.
                 items.append({"name": current_path or "List of items", "type": f"List (len {len(obj)})", "shape": "N/A", "object_path": current_path})
            else: # Inspect list items if they are containers
                for i, value in enumerate(obj):
                    new_path = f"{current_path}[{i}]" if current_path else f"item[{i}]"
                    # Limit recursion depth for lists to avoid excessively deep dives
                    if len(new_path.split(".")) < 5: # Arbitrary depth limit
                        items.extend(self._inspect_python_object(value, new_path))
        return items

    def inspect_pickle_hickle(self, file_path, file_type='pickle'):
        """
        Inspects a Pickle or Hickle file for loadable DataFrames or structures.
        Returns a list of discoverable datasets with their paths.
        """
        data_object = None
        try:
            if file_type == 'pickle':
                with open(file_path, 'rb') as f:
                    data_object = pickle.load(f)
            elif file_type == 'hickle':
                data_object = hickle.load(file_path)
            else:
                return [{"name": f"Unsupported type or library missing for {file_type}", "type": "Error", "shape": "N/A"}]

            return self._inspect_python_object(data_object)
        except Exception as e:
            print(f"Error inspecting {file_type} file {file_path}: {e}")
            return [{"name": f"Error: {e}", "type": "Error", "shape": "N/A"}]
            
    def _get_object_by_path(self, obj, path_str):
        """Retrieves a nested object using a path string like 'dict_key.list_index.attr'"""
        if not path_str: return obj # Root object
        
        current = obj
        parts = re.split(r'\.|\[(\d+)\]', path_str) # Split by . or [index]
        parts = [p for p in parts if p] # Remove empty strings from split

        for part in parts:
            if isinstance(current, dict):
                if part in current:
                    current = current[part]
                else: raise KeyError(f"Key '{part}' not found in dict.")
            elif isinstance(current, list):
                try:
                    idx = int(part)
                    current = current[idx]
                except (ValueError, IndexError):
                    raise ValueError(f"Invalid list index '{part}'.")
            # Add other types if needed (e.g., attributes of custom objects)
            else:
                raise TypeError(f"Cannot traverse object of type {type(current)} with path part '{part}'.")
        return current


    def load_selected_pickle_hickle_item(self, file_path, object_path, file_type='pickle'):
        """Loads a specific item (identified by its path) from a Pickle/Hickle file."""
        data_object = None
        try:
            if file_type == 'pickle':
                with open(file_path, 'rb') as f:
                    data_object = pickle.load(f)
            elif file_type == 'hickle':
                data_object = hickle.load(file_path)
            else:
                raise ValueError(f"Unsupported type or library missing for {file_type}")

            selected_item = self._get_object_by_path(data_object, object_path)

            if isinstance(selected_item, pl.DataFrame):
                return selected_item
            elif 'pandas' in sys.modules and isinstance(selected_item, sys.modules['pandas'].DataFrame):
                return pl.from_pandas(selected_item)
            elif isinstance(selected_item, np.ndarray):
                if selected_item.ndim == 1: return pl.DataFrame({object_path.split('.')[-1] or "data": selected_item})
                if selected_item.ndim == 2: return pl.DataFrame(selected_item)
                if selected_item.ndim == 0 and selected_item.dtype.fields: return pl.from_numpy(selected_item, orient="row")
            elif isinstance(selected_item, list) and selected_item:
                 if all(isinstance(row, dict) for row in selected_item): # List of dicts
                    return pl.DataFrame(selected_item)
            raise ValueError(f"Selected item at path '{object_path}' is not directly convertible to a Polars DataFrame.")
        except Exception as e:
            print(f"Error loading item '{object_path}' from {file_type} file {file_path}: {e}")
            raise


class FileLoaderWorker(QThread): # QThread is good if UI interaction is needed via signals
    request_excel_to_csv_conversion = Signal(str)
    request_npz_array_selection = Signal(list, str) # list of array_info, file_path
    npz_array_selected = Signal(str) # selected array name

    request_pickle_item_selection = Signal(list, str, str) # list of item_info, file_path, file_type
    pickle_item_selected = Signal(str) # selected item_path
    def __init__(self, file_path, sheet_name=None, usecols=None, header_row=None, 
                 dtype=None, data_only=True, read_only=True, nrows=None, run_wd=None, parent_gui=None): # Added parent_gui
        super().__init__()
        self.file_path = file_path
        self._file_path_ = file_path 
        self.sheet_name = sheet_name
        self.usecols = usecols
        self.header_row = header_row
        self.dtype_is_str = dtype == str # if True, cast all to string after loading
        self.nrows = nrows
        self.run_wd = run_wd
        self.npz_array_name = None # Can be set by QInputDialog if needed
        
        self.parent_gui = parent_gui # Reference to Pickaxe window for dialogs (or use signals)
        self.navigator = ContainerFileNavigator() # Instantiate navigator
        # For handling async selection from UI
        self.selected_npz_array_name = None
        self.selected_pickle_item_path = None
        self.selection_event = threading.Event()
        self.excel_conversion_event = threading.Event()
        self.user_wants_csv_conversion = False


    def wait_for_file_access(self, file_path, max_attempts=10, delay=1):
        for attempt in range(max_attempts):
            try:
                if not os.path.exists(file_path):
                    raise FileNotFoundError("Output file does not exist")
                if os.path.getsize(file_path) == 0:
                    raise ValueError("Output file is empty")
                with open(file_path, 'rb') as f:
                    f.read(4)
                return True
            except (PermissionError, FileNotFoundError, ValueError) as e:
                if attempt == max_attempts - 1:
                    raise Exception(f"File access failed after {max_attempts} attempts: {str(e)}")
                time.sleep(delay)
        return False


    def run(self, convert_all_to_csv=True):
        output_csv_path = None 
        df = None
        file_info = {'filename': self._file_path_} 
        start_time = time.time()

        # Use self._file_path_ for os.path operations if self.file_path might change (like in Excel conversion)
        # However, for reading, self.file_path should point to the actual file to be read.
        current_path_for_reading = self.file_path 
        file_ext = os.path.splitext(current_path_for_reading)[1].lower()

        try:            
            # Re-check file_ext as it might have changed if Excel was converted to CSV
            if file_ext == '.csv':
                # If df is already loaded (from Excel->CSV conversion), don't reload
                if df is None: 
                    df, csv_file_info = self.read_csv(current_path_for_reading)
                    file_info.update(csv_file_info)
                file_info['source_format'] = file_info.get('source_format', 'CSV') # Keep if already set by Excel conversion
            elif file_ext in ['.xlsx', '.xlsm', '.xls', '.xlsb']:
                df, excel_file_info = self.read_excel(current_path_for_reading, self.sheet_name, self.nrows)
                file_info.update(excel_file_info)
                file_info['source_format'] = 'Excel'
            elif file_ext == '.parquet':
                df = pl.read_parquet(current_path_for_reading, n_rows=self.nrows, use_pyarrow=False) 
                file_info['source_format'] = 'Parquet'
            elif file_ext == '.npz':
                df = self.handle_npz_load(current_path_for_reading)
                file_info['source_format'] = 'NumPy NPZ'
            elif file_ext in ['.pkl', '.pickle']:
                df = self.handle_pickle_hickle_load(current_path_for_reading, 'pickle')
                file_info['source_format'] = 'Pickle'
            elif file_ext in ['.hkl', '.hickle'] :
                df = self.handle_pickle_hickle_load(current_path_for_reading, 'hickle')
                file_info['source_format'] = 'Hickle'
            elif file_ext == '.json': 
                df = pl.read_json(current_path_for_reading)
                file_info['source_format'] = 'JSON'
            elif file_ext in ['.jsonl', '.ndjson']: 
                df = pl.read_ndjson(current_path_for_reading, n_rows=self.nrows)
                file_info['source_format'] = 'JSONL/NDJSON'
            elif file_ext in ['.dat', '.txt']:
                # Use the existing robust read_csv method
                df, txt_file_info = self.read_csv(current_path_for_reading) # Pass current_path_for_reading
                file_info.update(txt_file_info)
                file_info['source_format'] = f"Text ({file_ext})" # Be more specific
            else:
                # If we reach here and df is still None, it means an unhandled extension
                # or an Excel that wasn't converted (e.g., too small) but also wasn't handled by the Excel block
                if df is None: 
                    raise ValueError(f"Unsupported or unhandled file format: {file_ext}")

            if df is not None:
                if self.dtype_is_str: 
                    df = df.select([pl.all().cast(pl.Utf8)])
                file_info['load_time'] = time.time() - start_time
                file_info['rows'] = df.height
                file_info['columns'] = df.width
            
            if 'size' not in file_info and os.path.exists(self._file_path_): 
                 file_info['size'] = os.path.getsize(self._file_path_) / (1024 * 1024)


        except Exception as e:
            file_info['error'] = str(e) 
            print(f"Error in FileLoaderWorker.run for {self._file_path_}: {e}") 
        finally:
            if output_csv_path and os.path.exists(output_csv_path):
                try:
                    os.unlink(output_csv_path)
                except Exception as e_unlink:
                    print(f"Warning: Could not delete temporary CSV file {output_csv_path}: {e_unlink}")
        
        return df, file_info


    def read_excel(self, filename, sheet_name=None, nrows=None):
        try:
            # First attempt to read directly
            list_sheetnames = get_sheet_names(filename)
            opened_sheet = sheet_name if (sheet_name is not None and sheet_name in list_sheetnames and len(list_sheetnames) > 1) else list_sheetnames[0]
            
            df = pl.read_excel(
                source = filename,
                sheet_name=opened_sheet,
                infer_schema_length=0,
                # read_csv_options={'infer_schema_length': 0}
            )

            if nrows is not None and nrows > 0:
                df = df.slice(0, nrows)

            if self.dtype_is_str:
                df = df.select([pl.all().cast(pl.Utf8)])

            file_info = {
                'filename': filename, 'sheet_name': opened_sheet, 'total_sheets': len(list_sheetnames),
                'rows': df.height, 'columns': df.width, 'size': os.path.getsize(filename) / (1024 * 1024),
                'source_format': 'Excel'
            }
            return df, file_info
        
        except Exception as e:
            # If the direct read fails, ask the user about CSV conversion
            print(f"Direct Excel read failed: {e}. Requesting user input for CSV conversion.")
            self.excel_conversion_event.clear()
            self.request_excel_to_csv_conversion.emit(filename) # Signal to GUI
            self.excel_conversion_event.wait(timeout=60) # Wait for user's choice

            if hasattr(self, 'user_wants_csv_conversion') and self.user_wants_csv_conversion:
                # User said yes. Proceed with conversion.
                print("User approved CSV conversion. Attempting now...")
                try:
                    output_csv_path = Path(tempfile.gettempdir()).joinpath(
                        f"{Path(filename).stem.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    )
                    from xlsx2csv import Xlsx2csv
                    list_sheetnames_again = get_sheet_names(filename)
                    opened_sheet_again = sheet_name if sheet_name and sheet_name in list_sheetnames_again else list_sheetnames_again[0]
                    
                    converter = Xlsx2csv(filename, outputencoding="utf-8", delimiter=',')
                    converter.convert(str(output_csv_path), sheetname=opened_sheet_again)
                    
                    if self.wait_for_file_access(str(output_csv_path)):
                        df_csv, csv_file_info = self.read_csv(str(output_csv_path))
                        csv_file_info['source_format'] = 'Excel (converted to CSV after fail)'
                        csv_file_info['original_filename'] = filename
                        csv_file_info['sheet_name'] = opened_sheet_again
                        # Keep the file info pointing to the original excel file for user display
                        csv_file_info['filename'] = filename
                        csv_file_info['size'] = os.path.getsize(filename) / (1024*1024)
                        return df_csv, csv_file_info
                    else:
                        raise Exception("Conversion to CSV resulted in an empty or inaccessible file.")
                except Exception as conversion_error:
                    raise Exception(f"Initial read failed. CSV conversion also failed: {conversion_error}") from e
            else:
                # User said no or timed out. Re-raise the original exception.
                raise Exception("Failed to read Excel file directly and user declined CSV conversion attempt.") from e

    # Add this new slot method to the FileLoaderWorker class
    @Slot(bool)
    def on_excel_conversion_choice(self, proceed_with_conversion):
        self.user_wants_csv_conversion = proceed_with_conversion
        self.excel_conversion_event.set()


    def handle_npz_load(self, file_path):
        contents = self.navigator.inspect_npz(file_path)
        if not contents:
            raise ValueError("NPZ file is empty or unreadable.")
        if len(contents) == 1 and not contents[0]["name"].startswith("Error:"):
            self.selected_npz_array_name = contents[0]["name"]
        elif any(item["name"].startswith("Error:") for item in contents):
            raise ValueError(f"Error inspecting NPZ contents: {contents[0]['name']}")
        else:
            # Need to signal main thread to show a dialog
            self.selection_event.clear()
            self.request_npz_array_selection.emit(contents, os.path.basename(file_path))
            self.selection_event.wait(timeout=60) # Wait for user selection (with timeout)
            if not self.selected_npz_array_name:
                raise Exception("No array selected from NPZ file or selection timed out.")
        
        return self.read_npz(file_path, self.selected_npz_array_name)

    def handle_pickle_hickle_load(self, file_path, file_type):
        contents = self.navigator.inspect_pickle_hickle(file_path, file_type)
        if not contents:
            raise ValueError(f"{file_type.capitalize()} file appears empty or contains no recognized data structures.")
        
        loadable_items = [item for item in contents if not item["name"].startswith("Error:")]
        if not loadable_items:
             raise ValueError(f"No loadable datasets found in {file_type.capitalize()} file: {contents[0]['name'] if contents else 'Unknown error'}")

        if len(loadable_items) == 1:
            self.selected_pickle_item_path = loadable_items[0]["object_path"]
        else:
            self.selection_event.clear()
            self.request_pickle_item_selection.emit(loadable_items, os.path.basename(file_path), file_type)
            self.selection_event.wait(timeout=60)
            if not self.selected_pickle_item_path:
                raise Exception(f"No item selected from {file_type.capitalize()} file or selection timed out.")
        
        return self.navigator.load_selected_pickle_hickle_item(file_path, self.selected_pickle_item_path, file_type)

    # These slots would be connected to signals from the main GUI thread
    @Slot(str)
    def on_npz_array_selected(self, array_name):
        self.selected_npz_array_name = array_name
        self.selection_event.set()

    @Slot(str)
    def on_pickle_item_selected(self, item_path):
        self.selected_pickle_item_path = item_path
        self.selection_event.set()

    def detect_encoding_and_dialect(self, sampling_size_bytes=2**20):
        encoding = None
        with open(self.file_path, 'rb') as file:
            raw = file.read(128)
            for enc, boms in [('utf-8-sig', (codecs.BOM_UTF8,)),
                                ('utf-16', (codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)),
                                ('utf-32', (codecs.BOM_UTF32_LE, codecs.BOM_UTF32_BE))]:
                if any(raw.startswith(bom) for bom in boms):
                    encoding = enc
                    break
        if encoding is None:
            with open(self.file_path, 'rb') as f:
                result = detect(f.read(sampling_size_bytes))
            encoding = 'iso-8859-1' if result['encoding'] == 'ascii' else result['encoding']

        dialect = None
        try:
            with open(self.file_path, 'r', encoding=encoding, errors='ignore') as f:
                sample = f.read(sampling_size_bytes)
                if not sample:
                    dialect = excel; dialect.delimiter = ','; dialect.quotechar = '"'
                else:
                    sniffer_instance = Sniffer()
                    dialect = sniffer_instance.sniff(sample, delimiters=',;\t|')
                    if len(sample) > 100 and dialect.delimiter not in sample[:100]:
                        char_count = {char: sample.count(char) for char in [',', ';', '\t', '|'] if char in sample[:max(1000,len(sample)//10)]}
                        if char_count:
                            dialect.delimiter = max(char_count, key=lambda k: char_count[k])

        except Exception:
            dialect = excel
            dialect.delimiter = ','
            dialect.quotechar = '"'

        if not hasattr(dialect, 'delimiter') or not dialect.delimiter:
             dialect = excel
             dialect.delimiter = ','
             dialect.quotechar = '"'

        return encoding, dialect

    def read_csv(self, file_path):
        encoding, dialect = self.detect_encoding_and_dialect()
        df = None
        common_encodings = ['utf-8', 'iso-8859-1', 'windows-1252', 'utf-16']
        if encoding and encoding.lower() in common_encodings:
            common_encodings.remove(encoding.lower())

        attempt_encodings = ([encoding] if encoding else []) + common_encodings

        for enc_name in attempt_encodings:
            try:
                df = pl.read_csv(
                    file_path,
                    separator=dialect.delimiter,
                    quote_char=dialect.quotechar if dialect.quotechar else '"',
                    has_header=True,
                    encoding=enc_name if enc_name != 'utf-8-sig' else 'utf8',
                    infer_schema=False,
                    n_rows=self.nrows,
                    columns=self.usecols,
                    ignore_errors=True,
                    truncate_ragged_lines=True
                )
                if self.dtype_is_str:
                    df = df.select([pl.all().cast(pl.Utf8)])
                encoding = enc_name
                break
            except Exception as e:
                print(f"Failed to read CSV with encoding {enc_name} and delimiter '{dialect.delimiter}': {e}")
                continue

        if df is None:
            raise ValueError(f"Could not read CSV file {file_path} with attempted encodings/delimiters.")

        file_info = {
            'delimiter': dialect.delimiter,
            'encoding': encoding,
            'rows': df.height,
            'columns': df.width,
            'size': os.path.getsize(file_path) / (1024 * 1024),
            'dialect' : {
                'delimiter': dialect.delimiter,
                'quotechar': dialect.quotechar,
            }
        }
        return df, file_info

    def read_npz(self, file_path, array_name=None):
        with np.load(file_path) as data:
            keys = list(data.keys())
            if not keys:
                raise ValueError(".npz file contains no arrays.")
            
            # Simple case: load the first array if only one, or if a specific one is requested
            # A more complex UI would prompt the user to select an array from `keys`
            # For now, we'll try to be a bit smart or default to the first one.
            
            if array_name and array_name in keys:
                array_to_load_name = array_name
            else:
                array_to_load_name = keys[0] # Default to first array
                
            loaded_array = data[array_to_load_name]

            if loaded_array.ndim == 1:
                return pl.DataFrame(loaded_array)
            elif loaded_array.ndim == 2:
                # Try to create meaningful column names if possible
                col_names = [f"col_{i}" for i in range(loaded_array.shape[1])]
                return pl.DataFrame(loaded_array, schema=col_names)
            elif loaded_array.ndim == 0 and loaded_array.dtype.fields: # Structured array
                return pl.from_numpy(loaded_array, orient="row")
            else:
                raise ValueError(f"Array '{array_to_load_name}' in .npz file is not 1D or 2D, or a recognized structured array.")

    def read_pickle(self, file_path):
        # SECURITY WARNING: Pickle files can execute arbitrary code.
        # Only load from trusted sources. Consider adding a user warning.
        # For this application, we assume user trusts the source.
        with open(file_path, 'rb') as f:
            data_object = pickle.load(f)
        
        if isinstance(data_object, pl.DataFrame):
            return data_object
        elif 'pandas' in sys.modules and isinstance(data_object, sys.modules['pandas'].DataFrame):
            return pl.from_pandas(data_object)
        elif isinstance(data_object, np.ndarray):
            if data_object.ndim == 1: return pl.DataFrame({"column": data_object})
            if data_object.ndim == 2: return pl.DataFrame(data_object)
            if data_object.ndim == 0 and data_object.dtype.fields: return pl.from_numpy(data_object, orient="row")
        elif isinstance(data_object, list) and data_object:
            if all(isinstance(row, dict) for row in data_object):
                return pl.DataFrame(data_object)
        raise ValueError("Pickled object is not a Polars/Pandas DataFrame, NumPy array, or list of dicts.")

    def read_hickle(self, file_path):
        data_object = hickle.load(file_path)
        # Similar conversion logic as pickle
        if isinstance(data_object, pl.DataFrame):
            return data_object
        elif 'pandas' in sys.modules and isinstance(data_object, sys.modules['pandas'].DataFrame):
            return pl.from_pandas(data_object)
        elif isinstance(data_object, np.ndarray): # Add similar ndarray handling as pickle
            if data_object.ndim == 1: return pl.DataFrame({"column": data_object})
            if data_object.ndim == 2: return pl.DataFrame(data_object)
            if data_object.ndim == 0 and data_object.dtype.fields: return pl.from_numpy(data_object, orient="row")
        elif isinstance(data_object, list) and data_object:
            if all(isinstance(row, dict) for row in data_object):
                return pl.DataFrame(data_object)
        raise ValueError("Hickle object is not a Polars/Pandas DataFrame, NumPy array, or list of dicts.")


class FilledRowsBarWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.percentage = 0
        self.setMinimumHeight(20)
        self.setMinimumWidth(80)

    def set_fill_percentage(self, percentage):
        self.percentage = max(0, min(100, percentage))
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        width = self.width()
        height = self.height()
        painter.fillRect(0, 0, width, height, QColor(Qt.GlobalColor.lightGray))
        fill_color = QColor(Qt.GlobalColor.blue)
        if self.percentage <= 33: fill_color = QColor(Qt.GlobalColor.red)
        elif self.percentage <= 66: fill_color = QColor(255, 165, 0)
        else: fill_color = QColor(Qt.GlobalColor.green).darker(150)
        filled_width = int(width * self.percentage / 100.0)
        painter.fillRect(0, 0, filled_width, height, fill_color)
        text = f"{self.percentage:.1f}% filled"
        font = painter.font()
        font.setPointSize(max(8, int(height * 0.6)))
        painter.setFont(font)
        if fill_color == Qt.GlobalColor.red or \
           (fill_color == QColor(Qt.GlobalColor.green).darker(150) and self.percentage > 50) or \
           (fill_color == QColor(255,165,0) and self.percentage > 40) :
            if filled_width > painter.fontMetrics().horizontalAdvance(text) / 1.5 :
                painter.setPen(Qt.GlobalColor.white)
            else: painter.setPen(Qt.GlobalColor.black)
        else: painter.setPen(Qt.GlobalColor.black)
        painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, text)
        painter.end()

class FilterWorker(QThread):
    progress_updated = Signal(int, str)
    filter_completed = Signal(pl.DataFrame)

    def __init__(self, df, filters_and_logic):
        super().__init__()
        self.df = df
        self.filters_and_logic = filters_and_logic

    def run(self):
        current_df = self.df.clone()
        condition_list = []
        logic_list = []
        total_operations = len(self.filters_and_logic) # Should be sum of tuples
        operations_done = 0

        actual_filter_tuples = [item for item in self.filters_and_logic if isinstance(item, tuple)]
        total_filter_ops = len(actual_filter_tuples)


        for item_idx, item in enumerate(self.filters_and_logic):
            if isinstance(item, tuple):
                column, comparison, filter_string, filter_string2, case_sensitive, field_type, use_regex, negate_filter = item
                operations_done += 1
                self.progress_updated.emit(int((operations_done / total_filter_ops) * 100) if total_filter_ops > 0 else 0, 
                                           f"Applying filter {operations_done}/{total_filter_ops}...")

                current_condition = None
                if field_type == 'Date':
                    current_condition = self.apply_date_filter(column, comparison, filter_string, filter_string2)
                elif field_type == 'Numeric':
                    current_condition = self.apply_numeric_filter(column, comparison, filter_string, filter_string2)
                elif comparison in ["Is First Occurrence (Column)", 
                                    "Is Last Occurrence (Column)", 
                                    "Is Any Duplicate (Column)", # Changed from "Is Duplicate Value (Column)"
                                    "Is Only Unique (Column)"]: # Changed from "Is Unique Value (Column)"
                    current_condition = self.apply_column_occurrence_filter(column, comparison)
                else: # String filters
                    current_condition = self.apply_string_filter(column, comparison, filter_string, case_sensitive, use_regex)

                if current_condition is not None:
                    if negate_filter:
                        current_condition = ~current_condition
                    condition_list.append(current_condition)
                elif not (comparison in ["Is First Occurrence (Column)", "Is Last Occurrence (Column)", "Is Any Duplicate (Column)", "Is Only Unique (Column)"] and field_type == "String"): # Avoid error for non-applicable string options
                    pass # Or handle as an invalid filter combination if needed

            elif isinstance(item, str): # It's a logic string ("and" / "or")
                if condition_list: # Only add logic if there's a preceding condition
                     logic_list.append(item)

        if not condition_list: 
            filtered_df = self.df.clone() if not actual_filter_tuples else current_df # If no valid filters, return original or current state
            self.progress_updated.emit(100, "No effective filters applied." if not actual_filter_tuples else "Filtering completed.")
            self.filter_completed.emit(filtered_df)
            return
        
        final_condition = condition_list[0]
        for i in range(1, len(condition_list)):
            if i-1 < len(logic_list): # Ensure logic_list is not accessed out of bounds
                logic = logic_list[i-1]
                if logic.lower() == "or":
                    final_condition = final_condition | condition_list[i]
                else: # Default to AND
                    final_condition = final_condition & condition_list[i]
            else: # Should not happen with correct logic list building
                final_condition = final_condition & condition_list[i] 
        try:
            filtered_df = current_df.filter(final_condition)
        except Exception as e:
            print(f"Error during column filtering: {e}")
            self.progress_updated.emit(100, f"Error: {e}")
            self.filter_completed.emit(self.df.clear()) # Return empty on error
            return

        self.progress_updated.emit(100, "Filtering completed.")
        self.filter_completed.emit(filtered_df)

    def apply_date_filter(self, column, comparison, filter_string, filter_string2):
        try:
            if self.df.schema[column] == pl.Utf8:
                 col_expr = pl.col(column).str.to_date(format="%Y-%m-%d", strict=False)
            else:
                 col_expr = pl.col(column).cast(pl.Date, strict=False)

            if comparison == 'Empty': return pl.col(column).is_null()

            filter_date = datetime.strptime(filter_string, "%Y-%m-%d").date() if filter_string else None
            if filter_date is None: return pl.lit(False)

            if comparison in ('Equals', 'Contains'): return col_expr == filter_date
            elif comparison == '>': return col_expr > filter_date
            elif comparison == '<': return col_expr < filter_date
            elif comparison == '>=': return col_expr >= filter_date
            elif comparison == '<=': return col_expr <= filter_date
            elif comparison == 'Between':
                filter_date2 = datetime.strptime(filter_string2, "%Y-%m-%d").date() if filter_string2 else None
                if filter_date2 is None: return pl.lit(False)
                return (col_expr >= min(filter_date, filter_date2)) & (col_expr <= max(filter_date, filter_date2))
        except Exception as e:
            print(f"Date filter error for column '{column}': {e}")
            return pl.lit(False)
        return pl.lit(False)


    def apply_numeric_filter(self, column, comparison, filter_string, filter_string2):
        try:
            col_expr = pl.col(column).cast(pl.Float64, strict=False)
            if comparison == 'Empty': return pl.col(column).is_null()

            val1 = float(filter_string) if filter_string else None
            if val1 is None:
                 if comparison == 'Equals': return col_expr.is_null()
                 return pl.lit(False)

            if comparison == '>': return col_expr > val1
            elif comparison == '<': return col_expr < val1
            elif comparison == '>=': return col_expr >= val1
            elif comparison == '<=': return col_expr <= val1
            elif comparison == 'Equals': return col_expr == val1
            elif comparison == 'Between':
                val2 = float(filter_string2) if filter_string2 else None
                if val2 is None: return pl.lit(False)
                return (col_expr >= min(val1, val2)) & (col_expr <= max(val1, val2))
        except Exception as e:
            print(f"Numeric filter error for column '{column}': {e}")
            return pl.lit(False)
        return pl.lit(False)

    def apply_string_filter(self, column, comparison, filter_string, case_sensitive, use_regex):
        try:
            original_col_expr = pl.col(column)
            col_expr_str_casted = original_col_expr.cast(pl.Utf8).fill_null("")


            if comparison == 'Empty':
                return original_col_expr.is_null() | (col_expr_str_casted == "")

            fs = filter_string

            target_col_for_op = col_expr_str_casted
            fs_for_op = fs
            if not use_regex and not case_sensitive:
                target_col_for_op = col_expr_str_casted.str.to_lowercase()
                fs_for_op = fs.lower()

            if use_regex:
                pattern = fs
                if not case_sensitive:
                    pattern = f"(?i){pattern}"

                if comparison == 'Equals':
                    return target_col_for_op.str.contains(f"^{pattern}$")
                elif comparison == 'Contains':
                    return target_col_for_op.str.contains(pattern)
                else:
                    return pl.lit(False)
            else:
                if comparison == 'Contains':
                    return target_col_for_op.str.contains(fs_for_op, literal=True)
                elif comparison == 'Equals':
                    return target_col_for_op == fs_for_op
                elif comparison == '>': return target_col_for_op > fs_for_op
                elif comparison == '<': return target_col_for_op < fs_for_op
                elif comparison == '>=': return target_col_for_op >= fs_for_op
                elif comparison == '<=': return target_col_for_op <= fs_for_op
        except Exception as e:
            print(f"String filter error for column '{column}': {e}")
            return pl.lit(False)
        return pl.lit(False)

    def apply_column_duplicate_filter(self, column, comparison):
        try:
            if comparison == "Is Duplicate Value (Column)":
                return pl.col(column).is_duplicated()
            elif comparison == "Is Unique Value (Column)":
                return pl.col(column).is_unique()
        except Exception as e:
            print(f"Column duplicate filter error for column '{column}': {e}")
            return pl.lit(False)
        return pl.lit(False)

    def apply_column_occurrence_filter(self, column, comparison): # New name
        try:
            if comparison == "Is First Occurrence (Column)":
                return pl.col(column).is_first_distinct()
            elif comparison == "Is Last Occurrence (Column)":
                return pl.col(column).is_last_distinct()
            elif comparison == "Is Any Duplicate (Column)": # Renamed
                return pl.col(column).is_duplicated()
            elif comparison == "Is Only Unique (Column)": # Renamed
                return pl.col(column).is_unique()
        except Exception as e:
            print(f"Column occurrence filter error for column '{column}': {e}")
            return pl.lit(False) # Return a Polars expression that evaluates to False for all rows
        return pl.lit(False)


class FileSaverWorker(QThread):
    progress_updated = Signal(int, str)
    save_completed = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, df, file_name, progress_callback, completion_callback, error_callback, file_info_specs):
        super().__init__()
        self.df = df
        self.file_name = file_name
        self.progress_updated.connect(progress_callback)
        self.save_completed.connect(completion_callback)
        self.error_occurred.connect(error_callback)
        self.file_info_specs = file_info_specs

    def run(self):
        try:
            self.progress_updated.emit(25, f"Saving file '{self.file_name}'...")
            df_to_save = self.df
            if "__original_index__" in df_to_save.columns:
                df_to_save = df_to_save.drop("__original_index__")

            if self.file_name.lower().endswith('.csv'):
                dialect = self.file_info_specs.get('dialect')
                sep = dialect['delimiter'] if dialect and 'delimiter' in dialect else ','
                quote_char = dialect['quotechar'] if dialect and 'quotechar' in dialect else '"'
                df_to_save.write_csv(self.file_name, separator=sep, quote_char=quote_char, quote_style="non_numeric")
            elif self.file_name.lower().endswith('.xlsx'):
                import xlsxwriter
                df_to_save.write_excel(self.file_name, autofit=True)
            self.progress_updated.emit(100, "File saved successfully.")
            self.save_completed.emit(self.file_name)
        except Exception as e:
            self.error_occurred.emit(str(e))

class FieldTypeSwitch(QWidget):
    type_changed = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        self.combo = QComboBox()
        self.combo.addItems(['String', 'Date', 'Numeric'])
        layout.addWidget(self.combo)
        layout.setContentsMargins(0,0,0,0)
        self.combo.currentTextChanged.connect(self.type_changed.emit)

    def set_type(self, field_type): self.combo.setCurrentText(field_type)
    def get_type(self): return self.combo.currentText()


class FilterWidget(QFrame):
    filter_changed = Signal()

    def __init__(self, columns, parent=None):
        super().__init__(parent)
        self._layout = QHBoxLayout(self)
        fixed_width = 110
        self.column_combo = QComboBox()

        self.column_combo.addItems((columns if columns else []))
        self.column_combo.setFixedWidth(fixed_width + 20)


        self.field_type_switch = FieldTypeSwitch()
        self.field_type_switch.setFixedWidth(fixed_width -10)

        self.comparison_combo = QComboBox()
        # UPDATED COMPARISON OPTIONS
        self.string_comparisons = ['Contains', 'Equals', '>', '<', '>=', '<=', 'Between', 'Empty',
                                   'Is First Occurrence (Column)', 
                                   'Is Last Occurrence (Column)', 
                                   'Is Any Duplicate (Column)', 
                                   'Is Only Unique (Column)']
        self.numeric_date_comparisons = ['Equals', '>', '<', '>=', '<=', 'Between', 'Empty']
        
        # Initial population, will be updated by on_ui_changed
        self.comparison_combo.addItems(self.string_comparisons)
        self.comparison_combo.setFixedWidth(fixed_width + 30) # Adjusted width for longer items

        # ... (rest of FilterWidget __init__ - unchanged input fields, buttons, date edits) ...
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("Enter filter text...")
        self.filter_input2 = QLineEdit()
        self.filter_input2.setPlaceholderText("Upper bound for 'Between'")
        self.filter_input2.setVisible(False)

        self.case_sensitive_combo = QComboBox()
        self.case_sensitive_combo.addItems(['Case Insensitive', 'Case Sensitive'])
        self.case_sensitive_combo.setFixedWidth(fixed_width + 20)

        self.use_regex_checkbox = QCheckBox("Regex")
        self.regex_help_label = QLabel()
        self.regex_help_label.setTextFormat(Qt.TextFormat.RichText) # Allow HTML
        self.regex_help_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction) # Allow clicking links
        self.regex_help_label.setOpenExternalLinks(True) # Qt will handle opening the link
        self.regex_help_label.setText('<a href="https://www.regexone.com/">Regex Tutorial</a>')
        self.regex_help_label.setToolTip("Click to open regexone.com for a regex tutorial.")
        self.regex_help_label.setVisible(False) # Initially hidden

        self.negate_filter_checkbox = QCheckBox("Not")

        self.remove_button = QPushButton("X")
        self.remove_button.setFixedSize(30, 30)

        current_year = datetime.now().year
        default_date = QDate(current_year, 3, 31)
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDisplayFormat("yyyy-MM-dd")
        self.date_input.setDate(default_date)
        self.date_input.setVisible(False)
        self.date_input2 = QDateEdit()
        self.date_input2.setCalendarPopup(True)
        self.date_input2.setDisplayFormat("yyyy-MM-dd")
        self.date_input2.setDate(default_date)
        self.date_input2.setVisible(False)

        self._layout.addWidget(self.column_combo)
        self._layout.addWidget(self.field_type_switch)
        self._layout.addWidget(self.comparison_combo)
        self._layout.addWidget(self.filter_input)
        self._layout.addWidget(self.filter_input2)
        self._layout.addWidget(self.date_input)
        self._layout.addWidget(self.date_input2)
        self._layout.addWidget(self.case_sensitive_combo)
        self._layout.addWidget(self.use_regex_checkbox)
        self._layout.addWidget(self.regex_help_label) # Add the help label to the layout
        self._layout.addWidget(self.negate_filter_checkbox)
        self._layout.addWidget(self.remove_button)

        self.filter_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.date_input.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.date_input2.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

        self.column_combo.currentTextChanged.connect(self.on_column_or_comparison_changed)
        self.field_type_switch.type_changed.connect(self.on_ui_changed_and_emit)
        self.comparison_combo.currentTextChanged.connect(self.on_column_or_comparison_changed) # Re-check UI on comparison change too
        self.filter_input.textChanged.connect(self.on_ui_changed_and_emit)
        self.filter_input2.textChanged.connect(self.on_ui_changed_and_emit)
        self.date_input.dateChanged.connect(self.on_ui_changed_and_emit)
        self.date_input2.dateChanged.connect(self.on_ui_changed_and_emit)
        self.case_sensitive_combo.currentTextChanged.connect(self.on_ui_changed_and_emit)
        self.use_regex_checkbox.stateChanged.connect(self.on_ui_changed_and_emit)
        self.negate_filter_checkbox.stateChanged.connect(self.on_ui_changed_and_emit)
        
        self.on_ui_changed() # Initial UI setup

    def on_column_or_comparison_changed(self, text=None): 
        self.on_ui_changed() 
        self.filter_changed.emit() 

    def on_ui_changed_and_emit(self):
        self.on_ui_changed()
        self.filter_changed.emit()

    def on_ui_changed(self):
        field_type = self.field_type_switch.get_type()
        comparison = self.comparison_combo.currentText()
        
        # Update comparison combo items based on field type
        self.comparison_combo.blockSignals(True)
        current_comparison_text = self.comparison_combo.currentText()
        self.comparison_combo.clear()
        if field_type == 'String':
            self.comparison_combo.addItems(self.string_comparisons)
        else: # Numeric or Date
            self.comparison_combo.addItems(self.numeric_date_comparisons)
        
        # Try to restore previous selection
        idx = self.comparison_combo.findText(current_comparison_text)
        if idx != -1:
            self.comparison_combo.setCurrentIndex(idx)
        elif self.comparison_combo.count() > 0:
            self.comparison_combo.setCurrentIndex(0)
        self.comparison_combo.blockSignals(False)
        
        # Refresh comparison variable after list update
        comparison = self.comparison_combo.currentText() 
        is_regex_checked = self.use_regex_checkbox.isChecked()

        is_string_type = field_type == 'String'
        is_date_type = field_type == 'Date'
        is_numeric_type = field_type == 'Numeric'
        
        is_occurrence_filter = comparison in ["Is First Occurrence (Column)", 
                                              "Is Last Occurrence (Column)", 
                                              "Is Any Duplicate (Column)", 
                                              "Is Only Unique (Column)"]
        is_empty_comparison = comparison == 'Empty'

        show_filter_inputs = not is_empty_comparison and not is_occurrence_filter
        self.date_input.setVisible(is_date_type and show_filter_inputs)
        self.filter_input.setVisible(not is_date_type and show_filter_inputs)

        is_between = comparison == 'Between'
        self.date_input2.setVisible(is_date_type and is_between and show_filter_inputs)
        self.filter_input2.setVisible(not is_date_type and is_between and show_filter_inputs)

        can_show_regex_checkbox = is_string_type and comparison in ['Contains', 'Equals'] and show_filter_inputs
        self.use_regex_checkbox.setVisible(can_show_regex_checkbox)
        self.regex_help_label.setVisible(can_show_regex_checkbox) # Show/hide help label with checkbox

        if is_string_type and show_filter_inputs:
            self.filter_input.setPlaceholderText("Enter text (or regex if checked)" if can_show_regex_checkbox else "Enter filter text...")
        elif is_numeric_type and show_filter_inputs:
             self.filter_input.setPlaceholderText("Enter numeric value...")
        elif show_filter_inputs: 
            self.filter_input.setPlaceholderText("Enter filter text...")

        can_show_case_sensitive = is_string_type and comparison in ['Contains', 'Equals'] and not is_regex_checked and show_filter_inputs
        self.case_sensitive_combo.setVisible(can_show_case_sensitive)
        
        # Occurrence filters don't depend on field type for their direct operation, but field type switch should be disabled
        self.field_type_switch.setEnabled(not is_occurrence_filter)
        if is_occurrence_filter: # Force to string if it was an occurrence filter, as they make sense for any type
            self.field_type_switch.blockSignals(True)
            self.field_type_switch.set_type("String") # Or keep original type but ensure options are for column-wise ops
            self.field_type_switch.blockSignals(False)

    def get_filter_values(self):
        column = self.column_combo.currentText()
        comparison = self.comparison_combo.currentText()
        field_type = self.field_type_switch.get_type()
        filter_string = ""
        filter_string2 = ""

        is_value_input_needed = not (comparison == 'Empty' or \
                                   comparison in ["Is Duplicate Value (Column)", "Is Unique Value (Column)"])

        if is_value_input_needed:
            if field_type == 'Date':
                filter_string = self.date_input.date().toString("yyyy-MM-dd")
                filter_string2 = self.date_input2.date().toString("yyyy-MM-dd") if self.date_input2.isVisible() else ""
            else:
                filter_string = self.filter_input.text()
                filter_string2 = self.filter_input2.text() if self.filter_input2.isVisible() else ""

        case_sensitive = self.case_sensitive_combo.currentText() == 'Case Sensitive' if self.case_sensitive_combo.isVisible() else False
        use_regex = self.use_regex_checkbox.isChecked() if self.use_regex_checkbox.isVisible() else False
        negate_filter = self.negate_filter_checkbox.isChecked()

        actual_column = column


        return actual_column, comparison, filter_string, filter_string2, case_sensitive, field_type, use_regex, negate_filter

class LogicWidget(QFrame):
    logic_changed = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._layout = QHBoxLayout(self)
        self._layout.addStretch()
        self.and_radio = QRadioButton("AND")
        self.or_radio = QRadioButton("OR")
        self.and_radio.setChecked(True)
        self._layout.addWidget(self.and_radio)
        self._layout.addWidget(self.or_radio)
        self._layout.addStretch()
        self._layout.setContentsMargins(0,5,0,5)

        self.and_radio.toggled.connect(self.logic_changed.emit)


class FilterPanel(QWidget):
    panel_filters_changed = Signal()

    def __init__(self, columns, parent=None):
        super().__init__(parent)
        self._layout = QVBoxLayout(self)
        self.filters = []
        self.logic_widgets = []
        self.columns = columns if columns else []
        self.scroll_area = QScrollArea()
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        self.scroll_area.setWidget(self.scroll_widget)
        self.scroll_area.setWidgetResizable(True)
        self._layout.addWidget(self.scroll_area)
        self.button_layout = QHBoxLayout()
        self.add_filter_button = QPushButton("Add Filter Row")
        self.add_filter_button.clicked.connect(self.add_filter)
        self.apply_button = QPushButton("Apply Structured Filters")
        self.apply_button.setDefault(True)
        self.button_layout.addWidget(self.add_filter_button)
        self.button_layout.addWidget(self.apply_button)
        self._layout.addLayout(self.button_layout)
        if self.columns or True: 
            self.add_filter()
        else:
            self.add_filter_button.setEnabled(False)
            self.apply_button.setEnabled(False)

    def add_filter(self):
        filter_widget = FilterWidget(self.columns) # Pass current columns
        filter_widget.remove_button.clicked.connect(lambda checked=False, fw=filter_widget: self.remove_filter(fw))
        filter_widget.filter_changed.connect(self.panel_filters_changed.emit)

        if self.filters:
            logic_widget = LogicWidget()
            logic_widget.logic_changed.connect(self.panel_filters_changed.emit)
            self.logic_widgets.append(logic_widget)
            self.scroll_layout.addWidget(logic_widget)

        self.scroll_layout.addWidget(filter_widget)
        self.filters.append(filter_widget)
        self.panel_filters_changed.emit()

    def remove_filter(self, filter_widget_to_remove):
        try: index = self.filters.index(filter_widget_to_remove)
        except ValueError: return

        self.filters.pop(index)
        filter_widget_to_remove.filter_changed.disconnect(self.panel_filters_changed.emit)
        filter_widget_to_remove.deleteLater()

        if not self.filters and self.logic_widgets:
            for lw in self.logic_widgets:
                lw.logic_changed.disconnect(self.panel_filters_changed.emit)
                lw.deleteLater()
            self.logic_widgets.clear()
        elif index == 0 and self.logic_widgets:
            logic_widget_to_remove = self.logic_widgets.pop(0)
            logic_widget_to_remove.logic_changed.disconnect(self.panel_filters_changed.emit)
            logic_widget_to_remove.deleteLater()
        elif index > 0 and index <= len(self.logic_widgets):
            logic_widget_to_remove = self.logic_widgets.pop(index - 1)
            logic_widget_to_remove.logic_changed.disconnect(self.panel_filters_changed.emit)
            logic_widget_to_remove.deleteLater()

        self.apply_button.setEnabled(bool(self.filters))
        self.panel_filters_changed.emit()

    def get_filters_and_logic(self):
        filters_and_logic = []
        for i, filter_widget in enumerate(self.filters):
            if i > 0 and i <= len(self.logic_widgets):
                 filters_and_logic.append("and" if self.logic_widgets[i-1].and_radio.isChecked() else "or")
            filters_and_logic.append(filter_widget.get_filter_values())
        return filters_and_logic

class PolarsModel(QAbstractTableModel):
    def __init__(self, data_with_potential_index):
        super().__init__()
        self._full_data_with_index = data_with_potential_index
        self._data_for_cells = None
        self._display_columns = []
        self._original_indices = []

        if self._full_data_with_index is not None and not self._full_data_with_index.is_empty():
            self._row_count = self._full_data_with_index.height
            if "__original_index__" in self._full_data_with_index.columns:
                self._display_columns = [col for col in self._full_data_with_index.columns if col != "__original_index__"]
                self._data_for_cells = self._full_data_with_index.select(self._display_columns)
                self._original_indices = self._full_data_with_index.get_column("__original_index__").to_list()
            else: # Fallback if __original_index__ is not present
                self._display_columns = list(self._full_data_with_index.columns)
                self._data_for_cells = self._full_data_with_index
                self._original_indices = list(range(self._row_count))

            self._column_count = len(self._display_columns)
            self._column_headers = self._display_columns
            self._index_headers = [str(idx) for idx in self._original_indices]
        else:
            self._row_count = 0
            self._column_count = 0
            self._column_headers = []
            self._index_headers = []


    @property
    def data_frame(self): return self._data_for_cells # Returns only displayable data
    def rowCount(self, parent=None): return self._row_count
    def columnCount(self, parent=None): return self._column_count
    def data(self, index, role: int = Qt.ItemDataRole.DisplayRole):
        if self._data_for_cells is not None and index.isValid():
            if role == Qt.ItemDataRole.DisplayRole:
                try:
                    value = self._data_for_cells[index.row(), index.column()]
                    return str(value) if value is not None else ''
                except IndexError: return None
        return None

    @Slot(int, Qt.Orientation, int)
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                try: return str(self._column_headers[section])
                except IndexError: return None
            if orientation == Qt.Orientation.Vertical:
                try: return str(self._index_headers[section])
                except IndexError: return None
        return None

class AsyncFileLoaderThread(QThread):
    finished_signal = Signal(object, dict)
    error_signal = Signal(str)
    # New signals to relay from worker to main GUI for dialogs
    request_npz_array_selection = Signal(list, str)
    request_pickle_item_selection = Signal(list, str, str)
    request_excel_to_csv_conversion = Signal(str)

    def __init__(self, file_path, sheet_name=None, parent_gui=None, parent=None): # Added parent_gui
        super().__init__(parent)
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.parent_gui = parent_gui # Store parent_gui
        # The worker is created here, so it can be given the parent_gui if its methods need it
        # Or, connect worker's signals to this thread's signals
        self.worker_instance = FileLoaderWorker(file_path, sheet_name, parent_gui=parent_gui)
        
        # Relay signals from worker_instance to this thread's signals
        self.worker_instance.request_npz_array_selection.connect(self.request_npz_array_selection)
        self.worker_instance.request_pickle_item_selection.connect(self.request_pickle_item_selection)
        self.worker_instance.request_excel_to_csv_conversion.connect(self.request_excel_to_csv_conversion)
        
    @Slot(bool)
    def on_excel_conversion_choice_relayed(self, choice):
        if hasattr(self.worker_instance, 'on_excel_conversion_choice'):
            self.worker_instance.on_excel_conversion_choice(choice)

    def run(self):
        try:
            # FileLoaderWorker's run method is now directly called
            df, file_info = self.worker_instance.run()
            self.finished_signal.emit(df, file_info)
        except Exception as e:
            import traceback 
            self.error_signal.emit(f"Error loading file: {str(e)}\nTraceback:\n{traceback.format_exc()}")

    # Add slots to receive selection back from main GUI and pass to worker
    @Slot(str)
    def on_npz_array_selected_relayed(self, array_name):
        if hasattr(self.worker_instance, 'on_npz_array_selected'):
            self.worker_instance.on_npz_array_selected(array_name)

    @Slot(str)
    def on_pickle_item_selected_relayed(self, item_path):
        if hasattr(self.worker_instance, 'on_pickle_item_selected'):
            self.worker_instance.on_pickle_item_selected(item_path)

class Pickaxe(QMainWindow):
    update_model_signal = Signal(object, bool)
    visual_workshop_window = None
    data_transformer_window = None
    current_log_file_path = None
    types_suggested_and_applied_this_session = False
    
    def __init__(self):
        super().__init__()
        self.setAcceptDrops(True) # IMPORTANT: Enable drop events for the window

        self.setWindowTitle("Pickaxe")
        self.setGeometry(100, 100, 1200, 800)
        icon_path = resource_path("pickaxe.ico")
        if os.path.exists(icon_path): self.setWindowIcon(QIcon(icon_path))
        self.context_menu_column_logical_index = -1
        self.context_menu_column_name = None
        self.context_menu_selected_column_names = []
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self._layout = QVBoxLayout(self.central_widget)
        
        self.top_button_layout = QHBoxLayout()
        self.file_button = QPushButton("Open File")
        self.file_button.setDefault(True)
        self.file_button.setFixedWidth(self.file_button.size().width())
        self.file_button.clicked.connect(self.load_file)
        self.save_button = QPushButton("Save File")
        self.save_button.clicked.connect(self.save_file)
        
        # New Visual Workshop Button
        self.vw_button = QPushButton()
        self.vw_button.setFixedSize(30, 30) # Square button
        self.vw_button.setToolTip("Visualise Data")
        self.vw_button.clicked.connect(self.open_visual_workshop) 
        self.vw_button.setEnabled(False) # Enabled when data is loaded
        
        vw_icon_path = resource_path("vw.ico")
        if os.path.exists(vw_icon_path): 
            self.vw_button.setIcon(QIcon(vw_icon_path))
        else:
            self.vw_button = QPushButton()
            self._style_button(self.vw_button)

        # New "Suggest Type Conversions" button
        self.suggest_conversions_button = QPushButton()
        self.suggest_conversions_button.setFixedSize(30,30)
        self.suggest_conversions_button.setToolTip("Suggest Data Type Conversions")
        self.suggest_conversions_button.clicked.connect(self.suggest_and_convert_types)
        self.suggest_conversions_button.setEnabled(False)
        
        mw_icon_path = resource_path("magic-wand.ico")
        if os.path.exists(mw_icon_path): 
            self.suggest_conversions_button.setIcon(QIcon(mw_icon_path))
        else:
            self.suggest_conversions_button = QPushButton()
            self._style_button(self.suggest_conversions_button)
            
        # New Data Transformer Button
        self.dt_button = QPushButton()
        self.dt_button.setFixedSize(30, 30) # Square button
        self.dt_button.setToolTip("Transform Data")
        self.dt_button.clicked.connect(self.open_data_transformer)
        self.dt_button.setEnabled(False) # Enabled when data is loaded
        
        dt_icon_path = resource_path("settings.ico")
        if os.path.exists(dt_icon_path): 
            self.dt_button.setIcon(QIcon(dt_icon_path))
        else:
            self.dt_button = QPushButton()
            self._style_button(self.dt_button)

        self.top_button_layout.addWidget(self.file_button)
        self.top_button_layout.addWidget(self.save_button)
        self.top_button_layout.addWidget(self.dt_button)
        self.top_button_layout.addWidget(self.vw_button)
        self.top_button_layout.addWidget(self.suggest_conversions_button)
        # self.top_button_layout.addWidget(self.suggest_conversions_button)
        self._layout.addLayout(self.top_button_layout)
        self.info_layout = QHBoxLayout()
        self.file_info_label = QLabel("Load a file to see details.")
        self.file_info_label.setWordWrap(True)
        self.column_range_label = QLabel()
        self.info_layout.addWidget(self.file_info_label, 2)
        self.info_layout.addWidget(self.column_range_label, 1)
        self._layout.addLayout(self.info_layout)

        self.query_filter_layout = QHBoxLayout()
        self.query_input = QLineEdit()
        self.query_input.setPlaceholderText("Enter Polars filter expression (e.g., pl.col('Age') > 30 or pl.col('Name').str.contains(r'J[ao]ne? Doe'))") # Example with regex

        self.df = None

        self.completer = QCompleter(self)
        self.completer_model = QStringListModel(self)
        self.completer.setModel(self.completer_model)
        self.completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.completer.setFilterMode(Qt.MatchFlag.MatchContains)
        self.query_input.setCompleter(self.completer)
        self.update_completer_model()

        self.apply_query_button = QPushButton("Apply Query")
        self.apply_query_button.clicked.connect(self.apply_polars_query)
        
        self.query_help_button = QPushButton("?")
        self.query_help_button.setFixedSize(30, 30)
        self.query_help_button.setToolTip(
            "Open Polars expressions documentation.\n"
            "For string matching with regex, use methods like .str.contains(r'pattern').\n"
            "Learn regex at regexone.com (link also in structured filters)."
        )
        self.query_help_button.clicked.connect(self.show_query_help_extended) # New slot for extended help

        self.query_filter_layout.addWidget(self.query_input)
        self.query_filter_layout.addWidget(self.apply_query_button)
        self.query_filter_layout.addWidget(self.query_help_button)
        self._layout.addLayout(self.query_filter_layout)

        self.filter_panel = None
        self.filter_nav_layout = QHBoxLayout()
        self.filter_toggle_button = QPushButton("Show/Hide Structured Filters")
        self.filter_toggle_button.setFixedSize(self.file_button.size().width(), 30)
        self.filter_toggle_button.clicked.connect(self.toggle_filter_panel)
        self.filter_nav_layout.addWidget(self.filter_toggle_button)

        self.reset_filters_button = QPushButton("Reset All Filters/Sorting")
        self.reset_filters_button.clicked.connect(self.reset_all_filters)
        self.filter_nav_layout.addWidget(self.reset_filters_button)


        self.columns_per_page = 100
        self.nav_button_layout = QHBoxLayout()
        self.prev_columns_button = QPushButton(f"Previous {self.columns_per_page} Columns")
        self.prev_columns_button.clicked.connect(self.show_previous_columns)
        self.next_columns_button = QPushButton(f"Next {self.columns_per_page} Columns")
        self.next_columns_button.clicked.connect(self.show_next_columns)
        self.nav_button_layout.addWidget(self.prev_columns_button)
        self.nav_button_layout.addWidget(self.next_columns_button)
        self.filter_nav_layout.addLayout(self.nav_button_layout) # Add as a sub-layout
        self._layout.addLayout(self.filter_nav_layout)


        self.table_view = QTableView()
        self.table_view.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_view.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.show_context_menu)
        self._layout.addWidget(self.table_view)
        self.table_view.horizontalHeader().setSectionsClickable(True)
        self.table_view.horizontalHeader().setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table_view.horizontalHeader().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.horizontalHeader().customContextMenuRequested.connect(self.show_header_context_menu)
        self.stats_display_layout = QHBoxLayout()
        self.filled_rows_bar = FilledRowsBarWidget()
        self.filled_rows_bar.setFixedSize(120, 20)
        self.stats_display_layout.addWidget(self.filled_rows_bar, 0, Qt.AlignmentFlag.AlignVCenter)
        self.stats_label = QLabel("Select a column to see statistics.")
        self.stats_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.stats_label.setMinimumHeight(40)
        self.stats_label.setWordWrap(True)
        self.stats_label.setTextFormat(Qt.TextFormat.RichText) # <<< SET TO RICH TEXT
        self.stats_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction) # Optional: if you ever add links
        self.stats_label.setOpenExternalLinks(True) # Optional
        font = self.stats_label.font()
        font.setPointSize(9)
        self.stats_label.setFont(font)
        self.stats_display_layout.addWidget(self.stats_label, 1)
        self._layout.addLayout(self.stats_display_layout)
        self.model = None
        self._filepath = Path('')
        self.applied_filters_info = []
        self.file_info = {}
        self.current_column_page = 0
        self.filtered_df = None
        self.update_model_signal.connect(self.update_model_slot)
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setVisible(False)
        self.progress_label = QLabel()
        self.progress_label.setVisible(False)
        self._layout.addWidget(self.progress_bar)
        self._layout.addWidget(self.progress_label)
        self.save_button.setEnabled(False)
        self.apply_query_button.setEnabled(False)
        self.filter_toggle_button.setEnabled(False)
        self.reset_filters_button.setEnabled(False)
        self.prev_columns_button.setEnabled(False)
        self.next_columns_button.setEnabled(False)
        
        self.statusBar().show()
        
        # Default logger setup if no file loaded immediately (e.g. app starts empty)
        if not self.current_log_file_path and (self.df is not None) and (not self.df.is_empty()):
            default_log_path = os.path.join(os.getcwd(), f".__log__pickaxe_session_default_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md")
            logger.set_log_file(default_log_path, app_name_for_header="Pickaxe (No File Loaded)", associated_data_file="N/A")
            self.current_log_file_path = default_log_path


    def dragEnterEvent(self, event):
        # Called when a drag operation enters the widget
        mime_data = event.mimeData()
        if mime_data.hasUrls(): # Check if the drop data contains file paths
            # Check if any of the files are of acceptable types
            for url in mime_data.urls():
                file_path = url.toLocalFile()
                if os.path.isfile(file_path) and \
                   file_path.lower().endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.xlsb')):
                    event.acceptProposedAction() # Accept the drop if at least one is valid
                    return
        event.ignore() # Otherwise, ignore the drop

    def dragMoveEvent(self, event):
        # This can be used to give visual feedback during drag, but acceptProposedAction is often enough
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        # Called when the drop is completed
        mime_data = event.mimeData()
        if mime_data.hasUrls():
            file_paths_to_load = []
            for url in mime_data.urls():
                file_path = url.toLocalFile()
                if os.path.isfile(file_path) and \
                   file_path.lower().endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.xlsb')):
                    file_paths_to_load.append(file_path)
            
            if file_paths_to_load:
                # For simplicity, load the first valid file dropped.
                # You could extend this to handle multiple files (e.g., open multiple windows or tabs)
                self.statusBar().showMessage(f"File dropped: {os.path.basename(file_paths_to_load[0])}", 3000)
                self.load_file(file_paths_to_load[0]) 
                event.acceptProposedAction()
                return
        event.ignore()

    def _style_button(self, button):
        # Create a simple "logo" for the button
        pixmap = QPixmap(24, 24)
        pixmap.fill(Qt.GlobalColor.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        colors = [QColor("dodgerblue"), QColor("mediumseagreen"), QColor("tomato"), QColor("gold")]
        rect_size = 10
        offsets = [(0,0), (rect_size+2,0), (0, rect_size+2), (rect_size+2, rect_size+2)]
        for i in range(4):
            painter.setBrush(colors[i])
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(offsets[i][0], offsets[i][1], rect_size, rect_size, 2, 2)
        painter.end()
        button.setIcon(QIcon(pixmap))
        button.setIconSize(QSize(24,24))

    def reset_all_filters(self):
        if self.df is None:
            QMessageBox.information(self, "Reset Filters", "No data loaded to reset filters for.")
            return

        self.update_progress(0, "Resetting all filters...")
        QApplication.processEvents()

        self.filtered_df = self.df.clone() # Reset to original (with __original_index__)
        self.applied_filters_info = []
        self.query_input.clear()

        if self.filter_panel:
            # Simple reset: remove all filter rows and add one new one
            while self.filter_panel.filters:
                self.filter_panel.remove_filter(self.filter_panel.filters[0])
            if self.df is not None and not self.df.is_empty():
                 self.filter_panel.columns = [c for c in self.df.columns if c != "__original_index__"]
            else:
                self.filter_panel.columns = []
            self.filter_panel.add_filter() # Add a fresh filter row
            self.filter_panel.panel_filters_changed.emit() # To update query input if visible

        self.update_model_slot(self.filtered_df, True) # True to reset column page
        self.update_file_info_label() # Should reflect original counts now
        self.update_column_range_label(self.filtered_df)
        self.stats_label.setText("Select a column to see statistics.")
        self.filled_rows_bar.set_fill_percentage(0)
        self.set_stats_label_font_monospaced()
        self.update_progress(100, "All filters reset. Showing original data.")
        
        logger.log_action("Pickaxe", "Filters Reset", "All filters and sorting reset to original data.",
                            df_shape_after=self.df.shape if self.df is not None else None)



    def _get_df_to_operate_on(self):
        if self.filtered_df is not None and not self.filtered_df.is_empty():
            return self.filtered_df.clone()
        elif self.df is not None and not self.df.is_empty():
            return self.df.clone()
        return None

    def sort_by_column_action(self, ascending=True):
        if not self.context_menu_column_name:
            QMessageBox.warning(self, "Sort Error", "No column selected for sorting.")
            return

        col_name = self.context_menu_column_name
        df_to_modify = self._get_df_to_operate_on()

        if df_to_modify is None:
            QMessageBox.warning(self, "Sort Error", "No data to sort.")
            return

        self.update_progress(0, f"Sorting by '{col_name}'...")
        QApplication.processEvents()
        try:
            # Ensure __original_index__ is preserved if it exists
            cols_to_sort_by = [col_name]
            if "__original_index__" in df_to_modify.columns and col_name != "__original_index__":
                # Sorting might reorder, __original_index__ just comes along
                pass

            self.filtered_df = df_to_modify.sort(by=cols_to_sort_by, descending=(not ascending))
            self.applied_filters_info.append(f"Sort_{col_name}_{'ASC' if ascending else 'DESC'}")
            self.update_model_slot(self.filtered_df, True) # Reset page due to sort
            self.update_progress(100, "Sort completed.")
        except Exception as e:
            self.update_progress(100, "Sort failed.")
            QMessageBox.critical(self, "Sort Error", f"Error sorting by column '{col_name}':\n{e}")


    def remove_all_duplicate_rows_action(self, keep='first'):
        df_to_modify = self._get_df_to_operate_on()
        if df_to_modify is None:
            QMessageBox.warning(self, "Error", "No data to process.")
            return

        self.update_progress(0, f"Removing all duplicate rows (keep={keep})...")
        QApplication.processEvents()
        try:
            # Determine subset for uniqueness check (all columns except __original_index__)
            subset_cols = [col for col in df_to_modify.columns if col != "__original_index__"]
            if not subset_cols : # Should not happen if there's data
                self.filtered_df = df_to_modify
            else:
                self.filtered_df = df_to_modify.unique(subset=subset_cols, keep=keep, maintain_order=True)

            self.applied_filters_info.append(f"DedupAllRows_{keep[:1].upper()}")
            self.update_model_slot(self.filtered_df, True)
            self.update_progress(100, "All duplicate row removal completed.")
        except Exception as e:
            self.update_progress(100, "All duplicate row removal failed.")
            QMessageBox.critical(self, "Error", f"Error removing all duplicate rows:\n{e}")


    def update_completer_model(self):
        suggestions = []
        # Use self.df for columns as it's the base schema
        # but ensure __original_index__ is not suggested for direct manipulation by user
        if self.df is not None and not self.df.is_empty():
            display_cols = [c for c in self.df.columns if c != "__original_index__"]
            suggestions.extend([f'pl.col("{col}")' for col in display_cols])
            suggestions.extend([f'"{col}"' for col in display_cols])

        common_pl_starters = [
            "pl.col('')", "pl.lit()", "pl.when().then().otherwise()", "pl.sum('')", "pl.mean('')",
            "pl.count()", "pl.first()", "pl.last()", "pl.format('', )", "pl.all()", "pl.duration()",
            "pl.concat_str([], separator='')", "pl.datetime()", "pl.date()", "pl.time()"
        ]
        common_expr_methods = [
            ".cast(pl.)", ".is_in([])", ".is_null()", ".is_not_null()", ".alias('')",
            ".filter()", ".mean()", ".sum()", ".min()", ".max()", ".std()", ".var()", ".n_unique()",
            ".head()", ".tail()", ".sort()", ".value_counts()", ".unique()", ".is_duplicated()",
            ".is_first_distinct()", ".is_last_distinct()", ".fill_null()", ".drop_nulls()",
            ".round()", ".abs()", ".log()", ".exp()", ".pow()", ".gt()", ".lt()", ".eq()", ".neq()",
            ".is_between()", ".and_()", ".or_()", ".xor_()", ".is_nan()", ".is_not_nan()",
            ".is_finite()", ".is_infinite()", ".quantile()", ".agg()", ".over()"
        ]
        str_methods = [
            ".str.contains('')", ".str.starts_with('')", ".str.ends_with('')",
            ".str.to_lowercase()", ".str.to_uppercase()", ".str.strip_chars('')",
            ".str.replace_all('', '')", ".str.split('')", ".str.extract('')",
            ".str.len_bytes()", ".str.len_chars()", ".str.to_date()", ".str.to_datetime()",
            ".str.slice()", ".str.split_exact()", ".str.zfill()", ".str.pad_start()", ".str.pad_end()"
        ]
        dt_methods = [
            ".dt.year()", ".dt.month()", ".dt.day()", ".dt.hour()", ".dt.minute()",
            ".dt.second()", ".dt.microsecond()", ".dt.nanosecond()", ".dt.weekday()", ".dt.week()",
            ".dt.ordinal_day()", ".dt.strftime('')", ".dt.truncate('')", ".dt.offset_by('')",
            ".dt.round()", ".dt.replace_time_zone()", ".dt.combine()"
        ]
        pl_types = [
            "pl.Int8", "pl.Int16", "pl.Int32", "pl.Int64", "pl.UInt8", "pl.UInt16",
            "pl.UInt32", "pl.UInt64", "pl.Float32", "pl.Float64", "pl.Boolean", "pl.Utf8",
            "pl.Date", "pl.Datetime", "pl.Duration", "pl.Time", "pl.List", "pl.Struct"
        ]
        keywords_operators = ["True", "False", "None", "&", "|", "~", "==", "!=", ">", "<", ">=", "<="]

        suggestions.extend(common_pl_starters)
        suggestions.extend(common_expr_methods)
        suggestions.extend(str_methods)
        suggestions.extend(dt_methods)
        suggestions.extend(pl_types)
        suggestions.extend(keywords_operators)

        self.completer_model.setStringList(sorted(list(set(suggestions))))


    def generate_polars_expression_from_structured_filters(self):
        if not self.filter_panel or not self.filter_panel.filters:
            return ""

        filters_and_logic = self.filter_panel.get_filters_and_logic()
        if not filters_and_logic:
            return ""

        expression_parts = []

        for item_idx, item in enumerate(filters_and_logic):
            condition_str_part = ""
            if isinstance(item, tuple):
                col, comp, f1, f2, case, ftype, regex, negate = item

                if col is None:
                    continue # Skip if column is None for non-row-level operations

                col_expr_str = f'pl.col("{col}")'

                if comp == "Is Duplicate Value (Column)":
                    condition_str_part = f'{col_expr_str}.is_duplicated()'
                elif comp == "Is Unique Value (Column)":
                    condition_str_part = f'{col_expr_str}.is_unique()'
                elif ftype == "String":
                    base_col_str_op = f'{col_expr_str}.cast(pl.Utf8).fill_null("")'
                    target_col_str_op = base_col_str_op
                    single_backslash = r"\""
                    replacement_backslash = r"\\\""
                    val1_str_lit = f'"{f1.replace(single_backslash, replacement_backslash)}"'

                    if not regex and not case:
                        target_col_str_op = f'{base_col_str_op}.str.to_lowercase()'
                        val1_str_lit = f'"{f1.lower().replace(single_backslash, replacement_backslash)}"'

                    if comp == "Contains":
                        if regex:
                            pattern = f1 if case else f"(?i){f1}"
                            condition_str_part = f'{target_col_str_op}.str.contains(r"{pattern}")'
                        else:
                            condition_str_part = f'{target_col_str_op}.str.contains({val1_str_lit}, literal=True)'
                    elif comp == "Equals":
                        if regex:
                            pattern = f1 if case else f"(?i){f1}"
                            condition_str_part = f'{target_col_str_op}.str.contains(r"^{pattern}$")'
                        else:
                            condition_str_part = f'{target_col_str_op} == {val1_str_lit}'
                    elif comp == "Empty":
                        condition_str_part = f'(pl.col("{col}").is_null() | ({base_col_str_op} == ""))'
                    elif comp == ">": condition_str_part = f'{target_col_str_op} > {val1_str_lit}'
                    elif comp == "<": condition_str_part = f'{target_col_str_op} < {val1_str_lit}'
                    elif comp == ">=": condition_str_part = f'{target_col_str_op} >= {val1_str_lit}'
                    elif comp == "<=": condition_str_part = f'{target_col_str_op} <= {val1_str_lit}'


                elif ftype == "Numeric":
                    num_col_expr_str = f'{col_expr_str}.cast(pl.Float64, strict=False)'
                    try: val1_num = float(f1)
                    except ValueError: val1_num = None

                    if comp == "Empty": condition_str_part = f'pl.col("{col}").is_null()'
                    elif val1_num is not None:
                        if comp == "Equals": condition_str_part = f'{num_col_expr_str} == {val1_num}'
                        elif comp == ">": condition_str_part = f'{num_col_expr_str} > {val1_num}'
                        elif comp == "<": condition_str_part = f'{num_col_expr_str} < {val1_num}'
                        elif comp == ">=": condition_str_part = f'{num_col_expr_str} >= {val1_num}'
                        elif comp == "<=": condition_str_part = f'{num_col_expr_str} <= {val1_num}'
                        elif comp == "Between":
                            try: val2_num = float(f2)
                            except ValueError: val2_num = None
                            if val2_num is not None:
                                condition_str_part = f'({num_col_expr_str} >= {min(val1_num, val2_num)}) & ({num_col_expr_str} <= {max(val1_num, val2_num)})'
                            else: condition_str_part = "pl.lit(False)"
                        else: condition_str_part = "pl.lit(False)"
                    else: condition_str_part = "pl.lit(False)"

                elif ftype == "Date":
                    date_col_op_str = f'pl.col("{col}").str.to_date("%Y-%m-%d", strict=False, ambiguous="earliest").cast(pl.Date)'

                    if comp == "Empty": condition_str_part = f'pl.col("{col}").is_null()'
                    else:
                        try:
                            py_f1_date = datetime.strptime(f1, "%Y-%m-%d").date()
                            date_lit_f1 = f'pl.lit(datetime.date({py_f1_date.year}, {py_f1_date.month}, {py_f1_date.day})).cast(pl.Date)'

                            if comp == "Equals": condition_str_part = f'{date_col_op_str} == {date_lit_f1}'
                            elif comp == ">": condition_str_part = f'{date_col_op_str} > {date_lit_f1}'
                            elif comp == "<": condition_str_part = f'{date_col_op_str} < {date_lit_f1}'
                            elif comp == ">=": condition_str_part = f'{date_col_op_str} >= {date_lit_f1}'
                            elif comp == "<=": condition_str_part = f'{date_col_op_str} <= {date_lit_f1}'
                            elif comp == "Between":
                                py_f2_date = datetime.strptime(f2, "%Y-%m-%d").date()
                                min_py_date = min(py_f1_date, py_f2_date)
                                max_py_date = max(py_f1_date, py_f2_date)
                                date_lit_min = f'pl.lit(datetime.date({min_py_date.year}, {min_py_date.month}, {min_py_date.day})).cast(pl.Date)'
                                date_lit_max = f'pl.lit(datetime.date({max_py_date.year}, {max_py_date.month}, {max_py_date.day})).cast(pl.Date)'
                                condition_str_part = f'({date_col_op_str} >= {date_lit_min}) & ({date_col_op_str} <= {date_lit_max})'
                            else: condition_str_part = "pl.lit(False)"
                        except ValueError:
                            condition_str_part = "pl.lit(False)"

                if negate and condition_str_part and condition_str_part != "pl.lit(False)":
                    condition_str_part = f"~({condition_str_part})"


            if isinstance(item, tuple) and condition_str_part: # This condition means it's not a row-level filter
                if not expression_parts :
                    expression_parts.append(condition_str_part)
                else:
                    # Find the previous logic operator. Logic operators are strings.
                    # filters_and_logic could be [ (filter1), "and", (filter2), "or", (filter3_row_level_processed_to_none) ]
                    # We need to find the logic op that immediately preceded this filter tuple.
                    # It might not be item_idx-1 if some items were None.
                    
                    # Find the last valid logic operator string from the original list
                    # that appears before the current filter tuple.
                    last_logic_op_str = None
                    for k in range(item_idx -1, -1, -1):
                        if isinstance(filters_and_logic[k], str):
                            last_logic_op_str = filters_and_logic[k]
                            break
                    
                    if last_logic_op_str: # If a logic op was found and we have a previous expression part
                        op_symbol = "&" if last_logic_op_str.lower() == "and" else "|"
                        if expression_parts: # Ensure there's something to combine with
                            last_expr = expression_parts.pop()
                            if " & " in last_expr or " | " in last_expr: # Group if complex
                                last_expr = f"({last_expr})"
                            expression_parts.append(f"{last_expr} {op_symbol} {condition_str_part}")
                        else: # This is the first *evaluable* expression part after some non-evaluable ones.
                             expression_parts.append(condition_str_part)
                    else: # This must be the first *evaluable* filter
                        expression_parts.append(condition_str_part)


        final_expr_str = expression_parts[0] if expression_parts else ""
        return final_expr_str


    @Slot()
    def update_query_input_from_structured_filters(self):
        if self.filter_panel and self.filter_panel.isVisible():
            try:
                expression = self.generate_polars_expression_from_structured_filters()
                self.query_input.setText(expression)
            except Exception as e:
                print(f"Error generating Polars expression from structured filters: {e}")


    def handle_column_selection_for_stats(self, current, previous):
        if not current.isValid():
            self.stats_label.setText("No column selected or data not loaded.")
            self.filled_rows_bar.set_fill_percentage(0)
            self.set_stats_label_font_monospaced()
            return
        view_column_index = current.column()
        self.update_column_statistics_by_index(view_column_index)


    def update_column_statistics_by_index(self, view_column_index):
        if self.model is None or self.model.data_frame is None :
            self.stats_label.setText("No data loaded or displayed in table.")
            self.filled_rows_bar.set_fill_percentage(0)
            self.set_stats_label_font_monospaced() # Monospace font is good for text tables
            return

        active_df_for_stats = self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df
        if active_df_for_stats is None or active_df_for_stats.is_empty():
             self.stats_label.setText("No data to analyze.")
             self.filled_rows_bar.set_fill_percentage(0)
             self.set_stats_label_font_monospaced()
             return

        all_displayable_cols_in_active_df = [c for c in active_df_for_stats.columns if c != "__original_index__"]
        overall_display_col_idx = self.current_column_page * self.columns_per_page + view_column_index

        if not (0 <= overall_display_col_idx < len(all_displayable_cols_in_active_df)):
            self.stats_label.setText("Invalid column index for statistics.")
            self.filled_rows_bar.set_fill_percentage(0)
            self.set_stats_label_font_monospaced()
            return

        column_name = all_displayable_cols_in_active_df[overall_display_col_idx]
        series = active_df_for_stats.get_column(column_name)


        if series.is_empty():
            self.stats_label.setText(f"Column '{column_name}' is empty.")
            self.filled_rows_bar.set_fill_percentage(0)
            self.set_stats_label_font_monospaced()
            return

        not_na_count = series.len() - series.null_count()
        total_rows_in_series = series.len()
        fill_percentage = (not_na_count / total_rows_in_series) * 100 if total_rows_in_series > 0 else 0
        self.filled_rows_bar.set_fill_percentage(fill_percentage)

        stats_text_parts = []
        stats_text_parts.append(f"<b>Statistics for column: '{column_name}' (Type: {series.dtype})</b>")
        na_count = series.null_count()
        unique_count = series.n_unique()
        stats_text_parts.append(f"#Non-Empty: {not_na_count:,} | #Empty: {na_count:,} ({na_count / total_rows_in_series * 100:.1f}%) | #Unique: {unique_count:,}")

        # --- Unique Values Table using HTML ---
        stats_text_unique_values_html = ""
        max_unique_to_display = 10 # Always show top 10 if available

        if unique_count > 0:
            # Sort by count descending, then by value ascending as a tie-breaker
            value_counts_series_df = series.value_counts(sort=True).head(max_unique_to_display)
            
            # Start HTML table with some basic styling for borders
            stats_text_unique_values_html = "<br><br><b>Unique Values (Top {}):</b><br>".format(value_counts_series_df.height)
            stats_text_unique_values_html += "<table style='border-collapse: collapse; border: 1px solid #cccccc; width: auto;'>"
            
            # Header Row (Values)
            header_html = "<tr style='border-bottom: 1px solid #cccccc;'>"
            header_html += "<td style='border-right: 2px solid black; padding: 4px; font-weight: bold;'>Value</td>" # Bold first cell
            for i in range(value_counts_series_df.height):
                value = value_counts_series_df[column_name][i]
                str_value = str(value)
                # Truncate long values for display
                if len(str_value) > 12: # Adjust max length as needed
                    display_value = str_value[:10] + "..."
                else:
                    display_value = str_value
                # Italicize and center
                header_html += f"<td style='border-right: 1px solid #dddddd; padding: 4px; text-align: center;'><i>{display_value}</i></td>"
            header_html += "</tr>"
            stats_text_unique_values_html += header_html

            # Counts Row
            counts_html = "<tr>"
            counts_html += "<td style='border-right: 2px solid black; padding: 4px; font-weight: bold;'>Count(s)</td>" # Bold first cell
            for i in range(value_counts_series_df.height):
                count = value_counts_series_df["count"][i]
                counts_html += f"<td style='border-right: 1px solid #dddddd; padding: 4px; text-align: center;'>{count:,}</td>"
            counts_html += "</tr>"
            stats_text_unique_values_html += counts_html
            
            stats_text_unique_values_html += "</table>"

        elif unique_count == 0 and not_na_count > 0:
            stats_text_unique_values_html = "<br><br>No unique values (column might contain only one repeated value or all nulls)."
        elif not_na_count == 0: # All values are null
            stats_text_unique_values_html = "<br><br>All values in this column are empty/null."


        stats_text_parts.append(stats_text_unique_values_html)

        # --- Numeric Stats ---
        if series.dtype.is_numeric() and not_na_count > 0:
            # ... (your existing numeric stats formatting - unchanged) ...
            numeric_series_for_stats = series.drop_nulls().cast(pl.Float64)
            if not numeric_series_for_stats.is_empty():
                desc_df = numeric_series_for_stats.describe()
                desc = {row[0]: row[1] for row in desc_df.iter_rows()}
                sum_val = numeric_series_for_stats.sum()
                var_val = numeric_series_for_stats.var()
                zeros_count = numeric_series_for_stats.filter(numeric_series_for_stats == 0).len()
                pos_count = numeric_series_for_stats.filter(numeric_series_for_stats > 0).len()
                neg_count = numeric_series_for_stats.filter(numeric_series_for_stats < 0).len()
                stats_text_parts.append(f"<br><br><b>--- Numeric Stats ---</b>")
                stats_text_parts.append(f"Sum: {sum_val:,.2f} | Mean: {desc.get('mean', float('nan')):,.2f} | Median: {desc.get('median', float('nan')):,.2f}")
                stats_text_parts.append(f"Std Dev: {desc.get('std', float('nan')):,.2f} | Variance: {var_val:,.2f}")
                stats_text_parts.append(f"Min: {desc.get('min', float('nan')):,.2f} | 25% (Q1): {desc.get('25%', float('nan')):,.2f} | 75% (Q3): {desc.get('75%', float('nan')):,.2f} | Max: {desc.get('max', float('nan')):,.2f}")
                stats_text_parts.append(f"Counts: Zeros: {zeros_count:,} | Positive: {pos_count:,} | Negative: {neg_count:,}")

        # --- Datetime Stats ---
        elif (series.dtype == pl.Datetime or series.dtype == pl.Date) and not_na_count > 0:
            # ... (your existing datetime stats formatting - unchanged) ...
            datetime_series_for_stats = series.drop_nulls()
            if not datetime_series_for_stats.is_empty():
                stats_text_parts.append(f"<br><br><b>--- Datetime Stats ---</b>")
                min_date_val = datetime_series_for_stats.min()
                max_date_val = datetime_series_for_stats.max()
                stats_text_parts.append(f"Earliest: {min_date_val} | Latest: {max_date_val}")

                if min_date_val is not None and max_date_val is not None and max_date_val >= min_date_val:
                    py_min_date = min_date_val
                    py_max_date = max_date_val
                    if isinstance(min_date_val, date) and not isinstance(min_date_val, datetime):
                        py_min_date = datetime(min_date_val.year, min_date_val.month, min_date_val.day)
                    if isinstance(max_date_val, date) and not isinstance(max_date_val, datetime):
                        py_max_date = datetime(max_date_val.year, max_date_val.month, max_date_val.day)

                    if py_min_date and py_max_date:
                        time_span_delta = py_max_date - py_min_date
                        days_total = time_span_delta.days
                        years_span, months_span, days_span = days_total // 365, (days_total % 365) // 30, (days_total % 365) % 30
                        time_span_str_parts = []
                        if years_span > 0: time_span_str_parts.append(f"{years_span} year{'s' if years_span > 1 else ''}")
                        if months_span > 0: time_span_str_parts.append(f"{months_span} month{'s' if months_span > 1 else ''}")
                        if days_span > 0 or not time_span_str_parts : time_span_str_parts.append(f"{days_span} day{'s' if days_span > 1 else ''}")
                        time_span_str = ", ".join(time_span_str_parts) if time_span_str_parts else "0 days"
                        if days_total > 0 : time_span_str += f" [{days_total:,} total day{'s' if days_total > 1 else ''}]"
                        stats_text_parts.append(f"Time Span: {time_span_str}")
        
        # Join all parts. Use <br> for newlines in HTML.
        # The stats_label needs to support RichText.
        self.stats_label.setTextFormat(Qt.TextFormat.RichText)
        self.stats_label.setText("<br>".join(stats_text_parts))
        self.set_stats_label_font_monospaced() # Monospace is good for table-like text
                                
    def set_stats_label_font_monospaced(self):
        font = QFont()
        families = ["Courier New", "Consolas", "DejaVu Sans Mono", "Menlo"]
        for family in families:
            font.setFamily(family)
            if QFont(font.family()).styleHint() == QFont.StyleHint.Monospace: break
        else: font.setStyleHint(QFont.StyleHint.Monospace)
        font.setPointSize(9)
        self.stats_label.setFont(font)

    def show_header_context_menu(self, position):
        header = self.table_view.horizontalHeader()
        clicked_logical_view_index = header.logicalIndexAt(position) # This is view index on current page
        selected_model_indices = header.selectionModel().selectedIndexes()
        # These are also view indexes on current page
        selected_view_indexes_on_page = sorted(list(set(idx.column() for idx in selected_model_indices)))


        df_for_columns = self._get_df_to_operate_on() # Gets filtered or original df
        if df_for_columns is None: return

        all_displayable_cols = [c for c in df_for_columns.columns if c != "__original_index__"]

        self.context_menu_selected_column_names = []
        for view_idx_on_page in selected_view_indexes_on_page:
            overall_display_col_idx = self.current_column_page * self.columns_per_page + view_idx_on_page
            if 0 <= overall_display_col_idx < len(all_displayable_cols):
                self.context_menu_selected_column_names.append(all_displayable_cols[overall_display_col_idx])

        self.context_menu_column_name = None # This is the column name for single-column operations
        if clicked_logical_view_index != -1: # A specific column header was right-clicked
            overall_clicked_display_idx = self.current_column_page * self.columns_per_page + clicked_logical_view_index
            if 0 <= overall_clicked_display_idx < len(all_displayable_cols):
                self.context_menu_column_name = all_displayable_cols[overall_clicked_display_idx]
        elif self.context_menu_selected_column_names: # Fallback to first selected if multiple are selected but no specific right-click
             self.context_menu_column_name = self.context_menu_selected_column_names[0]


        menu = QMenu()

        if self.context_menu_column_name: # Operations requiring a single target column
            sort_menu = QMenu(f"Sort by '{self.context_menu_column_name}'", self)
            sort_asc_action = QAction("Ascending (A-Z, 0-9)", self)
            sort_asc_action.triggered.connect(lambda: self.sort_by_column_action(ascending=True))
            sort_menu.addAction(sort_asc_action)

            sort_desc_action = QAction("Descending (Z-A, 9-0)", self)
            sort_desc_action.triggered.connect(lambda: self.sort_by_column_action(ascending=False))
            sort_menu.addAction(sort_desc_action)
            menu.addMenu(sort_menu)

        # "Filter Duplicate Rows (Entire Row)" operations - these don't depend on a specific column click
        duplicates_all_rows_menu = QMenu("Filter Duplicate Rows (Entire Row)", self)
        remove_all_dup_rows_first_action = QAction("Keep First Occurrence", self)
        remove_all_dup_rows_first_action.triggered.connect(lambda: self.remove_all_duplicate_rows_action(keep='first'))
        duplicates_all_rows_menu.addAction(remove_all_dup_rows_first_action)

        remove_all_dup_rows_last_action = QAction("Keep Last Occurrence", self)
        remove_all_dup_rows_last_action.triggered.connect(lambda: self.remove_all_duplicate_rows_action(keep='last'))
        duplicates_all_rows_menu.addAction(remove_all_dup_rows_last_action)

        remove_all_dup_rows_none_action = QAction("Remove All Occurrences of Duplicates", self) # effectively keep unique
        remove_all_dup_rows_none_action.triggered.connect(lambda: self.remove_all_duplicate_rows_action(keep='none'))
        duplicates_all_rows_menu.addAction(remove_all_dup_rows_none_action)
        menu.addMenu(duplicates_all_rows_menu)


        if self.context_menu_selected_column_names: # Batch conversion applies to selected columns
            menu.addSeparator()
            num_sel = len(self.context_menu_selected_column_names)
            # For display in menu action, use the right-clicked column if it's among selected, else first selected
            col_name_display_for_convert = self.context_menu_column_name
            if self.context_menu_column_name not in self.context_menu_selected_column_names and self.context_menu_selected_column_names:
                col_name_display_for_convert = self.context_menu_selected_column_names[0]


            convert_numeric_text = f"Convert '{col_name_display_for_convert}' to Numeric" if num_sel == 1 else f"Convert {num_sel} Selected to Numeric"
            convert_datetime_text = f"Convert '{col_name_display_for_convert}' to Datetime" if num_sel == 1 else f"Convert {num_sel} Selected to Datetime"

            convert_numeric_action = QAction(convert_numeric_text, self)
            convert_numeric_action.triggered.connect(self.convert_selected_columns_to_numeric_batch if num_sel > 1 else self.convert_column_to_numeric)
            convert_datetime_action = QAction(convert_datetime_text, self)
            convert_datetime_action.triggered.connect(self.convert_selected_columns_to_datetime_batch if num_sel > 1 else self.convert_column_to_datetime)

            menu.addAction(convert_numeric_action)
            menu.addAction(convert_datetime_action)

        if menu.actions():
            menu.exec(header.mapToGlobal(position))


    def _convert_column_type(self, conversion_expression_constructor, type_name, target_polars_type=None): # Added target_polars_type
        if self.df is None or self.context_menu_column_name is None:
            QMessageBox.warning(self, "Conversion Error", "No data or column selected for conversion.")
            return
        column_name_to_convert = self.context_menu_column_name
        df_shape_before_conversion = self.df.shape
        original_column_series = self.df.get_column(column_name_to_convert) # Get before modification for logging

        try:
            self.progress_bar.setVisible(True)
            self.update_progress(0, f"Converting column '{column_name_to_convert}' to {type_name}...")
            QApplication.processEvents()

            original_column_series = self.df.get_column(column_name_to_convert)
            original_nulls_df = original_column_series.is_null().sum()

            the_conversion_expr = conversion_expression_constructor(pl.col(column_name_to_convert))

            temp_converted_series = self.df.select(the_conversion_expr.alias("___temp_conversion_check___")).get_column("___temp_conversion_check___")
            converted_nulls_temp = temp_converted_series.is_null().sum()

            significant_new_nans = False
            if converted_nulls_temp > original_nulls_df + (0.1 * self.df.height):
                if type_name == "datetime" and temp_converted_series.dtype in [pl.Datetime, pl.Date]:
                     pass
                else:
                    significant_new_nans = True

            if significant_new_nans:
                reply = QMessageBox.question(self, "Conversion Warning",
                                             f"Converting '{column_name_to_convert}' to {type_name} resulted in a significant number of new empty values. Proceed?",
                                             QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.No:
                    self.update_progress(0, "Conversion cancelled.")
                    self.progress_bar.setVisible(False); self.progress_label.setVisible(False)
                    return

            self.df = self.df.with_columns(the_conversion_expr.alias(column_name_to_convert))
            if self.filtered_df is not None and column_name_to_convert in self.filtered_df.columns:
                 self.filtered_df = self.filtered_df.with_columns(the_conversion_expr.alias(column_name_to_convert))

            self.update_progress(50, "Updating view...")
            QApplication.processEvents()
            self.update_model_slot(self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df, False)


            if hasattr(self, 'stats_label') and self.table_view.selectionModel().hasSelection():
                 current_selection = self.table_view.selectionModel().currentIndex()
                 if current_selection.isValid(): self.handle_column_selection_for_stats(current_selection, None)
            self.update_progress(100, f"Column '{column_name_to_convert}' converted to {type_name}.")
            
            logger.log_action(
                "Pickaxe", "Type Conversion", f"Column '{column_name_to_convert}' to {type_name}.",
                details={"Column": column_name_to_convert, 
                         "New Type": type_name, 
                         "Target Polars Type": str(target_polars_type) if target_polars_type else "N/A",
                         "Old Type": str(original_column_series.dtype)},
                df_shape_before=df_shape_before_conversion, 
                df_shape_after=self.df.shape
            )
            self.types_suggested_and_applied_this_session = True # Mark that types have been modified

        except Exception as e:
            self.update_progress(100, "Conversion failed.")
            QMessageBox.critical(self, "Conversion Error", f"Could not convert column '{column_name_to_convert}' to {type_name}:\n{e}")
            logger.log_action(
                "Pickaxe", "Type Conversion Error", f"Failed converting '{column_name_to_convert}' to {type_name}.",
                details={"Error": str(e)}
            )
        finally:
            self.progress_bar.setVisible(False); self.progress_label.setVisible(False)

    def convert_column_to_numeric(self):
        self._convert_column_type(lambda c_expr: c_expr.cast(pl.Float64, strict=False), "numeric")

    def convert_column_to_datetime(self):
        if self.df is None or self.context_menu_column_name is None:
            QMessageBox.warning(self, "Conversion Error", "No data or column selected for conversion.")
            return
        col_name = self.context_menu_column_name
        current_dtype_in_df = self.df.schema[col_name]

        if current_dtype_in_df == pl.Utf8:
            expr_builder = lambda c_expr: c_expr.str.to_datetime(strict=False, time_unit='ns', ambiguous='earliest').dt.replace_time_zone(None)
        else:
            expr_builder = lambda c_expr: c_expr.cast(pl.Datetime, strict=False).dt.replace_time_zone(None)

        self._convert_column_type(expr_builder, "datetime")


    def _batch_convert_columns(self, type_name, target_polars_type=None):
        if not hasattr(self, 'context_menu_selected_column_names') or not self.context_menu_selected_column_names:
            QMessageBox.warning(self, "Batch Conversion Error", "No columns selected for batch conversion.")
            return
        selected_names = self.context_menu_selected_column_names
        num_selected = len(selected_names)
        if num_selected == 0: return

        self.progress_bar.setVisible(True); self.progress_label.setVisible(True)
        converted_count = 0
        skipped_details = []
        error_details = []
        user_cancelled_all = False
        yes_to_all_warnings = False
        df_shape_before_batch_op = 0
        
        if self.df is not None:
            
            df_shape_before_batch_op = self.df.shape if not (self.df.is_empty()) else None

            for i, col_name in enumerate(selected_names):
                self.update_progress(int(((i + 0.5) / num_selected) * 100), f"Converting {i+1}/{num_selected}: '{col_name}' to {type_name}...")
                QApplication.processEvents()

                try:
                    if col_name not in self.df.columns:
                        error_details.append(f"'{col_name}': Column not found in DataFrame.")
                        continue

                    original_column_series = self.df.get_column(col_name)
                    original_nulls = original_column_series.is_null().sum()
                    current_col_schema_dtype = self.df.schema[col_name]

                    final_conversion_expr = None
                    actual_target_polars_type = None 

                    if type_name == "numeric_float": 
                        actual_target_polars_type = pl.Float64
                        final_conversion_expr = pl.col(col_name).cast(pl.Float64, strict=False)
                    elif type_name == "integer": 
                        actual_target_polars_type = pl.Int64
                        final_conversion_expr = pl.col(col_name).cast(pl.Int64, strict=False)
                    elif type_name == "datetime": 
                        actual_target_polars_type = pl.Datetime
                        if current_col_schema_dtype == pl.Utf8:
                            final_conversion_expr = pl.col(col_name).str.to_datetime(strict=False, time_unit='ns', ambiguous='earliest', format=None).dt.replace_time_zone(None)
                        else: 
                            final_conversion_expr = pl.col(col_name).cast(pl.Datetime, strict=False).dt.replace_time_zone(None)
                    elif type_name == "categorical": 
                        actual_target_polars_type = pl.Categorical
                        final_conversion_expr = pl.col(col_name).cast(pl.Categorical, strict=False) # Use strict=False initially
                    elif type_name == "boolean": # New Boolean type
                        actual_target_polars_type = pl.Boolean
                        # Polars' cast to Boolean is quite good with "true"/"false", 0/1
                        # For more complex strings ("Yes", "No"), pre-processing might be needed
                        # but for now, a direct cast is a good first step.
                        # Map common string representations of booleans to actual booleans before casting
                        if current_col_schema_dtype == pl.Utf8:
                            col_lower = pl.col(col_name).str.to_lowercase()
                            final_conversion_expr = (
                                pl.when(col_lower.is_in(["true", "yes", "1", "t"]))
                                .then(pl.lit(True, dtype=pl.Boolean))
                                .when(col_lower.is_in(["false", "no", "0", "f"]))
                                .then(pl.lit(False, dtype=pl.Boolean))
                                .otherwise(pl.lit(None, dtype=pl.Boolean)) # Map others to Null
                            )
                        else: # If not string, direct cast
                            final_conversion_expr = pl.col(col_name).cast(pl.Boolean, strict=False)

                    elif type_name == "string_key": 
                        actual_target_polars_type = pl.Utf8
                        final_conversion_expr = pl.col(col_name).cast(pl.Utf8)
                    elif type_name == "numeric": # Context menu "Convert to Numeric" usually implies float
                        actual_target_polars_type = pl.Float64
                        final_conversion_expr = pl.col(col_name).cast(pl.Float64, strict=False)


                    if final_conversion_expr is None:
                        error_details.append(f"'{col_name}': No conversion logic for target type name '{type_name}'.")
                        continue
                
                    temp_converted_series = self.df.select(final_conversion_expr.alias("___temp_batch_check___")).get_column("___temp_batch_check___")
                    converted_nulls = temp_converted_series.is_null().sum()
                    newly_created_nulls = converted_nulls - original_nulls

                    significant_new_nans = False
                    original_non_null_count = self.df.height - original_nulls
                    if original_non_null_count > 0 and (newly_created_nulls / original_non_null_count > 0.1):
                        significant_new_nans = True
                    elif newly_created_nulls > 0.05 * self.df.height:
                        significant_new_nans = True
                    
                    if type_name == "string_key" or type_name == "boolean": # Don't usually warn for these if result is mostly nulls as intended
                        significant_new_nans = False

                    if significant_new_nans and not yes_to_all_warnings:
                        msg_box = QMessageBox(self); msg_box.setIcon(QMessageBox.Icon.Warning)
                        msg_box.setWindowTitle(f"Conversion Warning for '{col_name}'")
                        msg_box.setText(f"Converting '{col_name}' to {type_name} resulted in {newly_created_nulls} new empty/null values. Proceed with this column?")
                        yes_button = msg_box.addButton("Yes", QMessageBox.ButtonRole.YesRole); no_button = msg_box.addButton("No", QMessageBox.ButtonRole.NoRole)
                        yes_all_button = msg_box.addButton("Yes to All", QMessageBox.ButtonRole.AcceptRole); cancel_all_button = msg_box.addButton("Cancel All", QMessageBox.ButtonRole.RejectRole)
                        msg_box.setDefaultButton(yes_button); msg_box.exec()
                        
                        clicked_btn = msg_box.clickedButton()
                        if clicked_btn == no_button: 
                            skipped_details.append(f"'{col_name}': Skipped by user due to new nulls."); continue
                        elif clicked_btn == cancel_all_button: 
                            user_cancelled_all = True; skipped_details.append(f"'{col_name}': Batch cancelled by user."); break
                        elif clicked_btn == yes_all_button: 
                            yes_to_all_warnings = True


                    self.df = self.df.with_columns(final_conversion_expr.alias(col_name))
                    if self.filtered_df is not None and col_name in self.filtered_df.columns:
                        self.filtered_df = self.filtered_df.with_columns(final_conversion_expr.alias(col_name))
                    converted_count += 1
                        
                except Exception as e: 
                    error_details.append(f"'{col_name}': {str(e)}")
                    print(f"Error converting {col_name} to {type_name}: {e}") 
                    import traceback
                    traceback.print_exc()


        self.update_progress(95, "Updating view after batch conversion...")
        QApplication.processEvents()
        self.update_model_slot(self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df, False) 
        
        self.update_completer_model() 
        
        if hasattr(self, 'stats_label') and self.table_view.selectionModel().hasSelection():
            current_selection = self.table_view.selectionModel().currentIndex()
            if current_selection.isValid(): self.handle_column_selection_for_stats(current_selection, None)

        summary_parts = [f"{converted_count} of {num_selected} columns processed for conversion to '{type_name}'." if not user_cancelled_all else f"Batch conversion to '{type_name}' cancelled. {converted_count} processed."]
        if skipped_details: summary_parts.append("\nSkipped:\n" + "\n".join(skipped_details))
        if error_details: summary_parts.append("\nErrors:\n" + "\n".join(error_details))
        
        if skipped_details or error_details or converted_count < num_selected : 
            QMessageBox.information(self, "Batch Conversion Summary", "\n".join(summary_parts))
        
        self.update_progress(100, f"Batch {type_name} conversion finished.")
        self.types_suggested_and_applied_this_session = True

        if self.df is not None:
            df_shape_after_batch_op = self.df.shape if not self.df.is_empty() else None
        else:
            df_shape_after_batch_op = 0
            
        logger.log_action("Pickaxe", "Batch Type Conversion", 
                        f"Attempted to convert {num_selected} columns to {type_name}.",
                        details={"Columns Selected": selected_names, 
                                "Target Type Name": type_name,
                                "Target Polars Type": str(target_polars_type if target_polars_type else "Inferred from type_name"),
                                "Converted Count": converted_count,
                                "Skipped": skipped_details, "Errors": error_details},
                        df_shape_before=df_shape_before_batch_op,
                        df_shape_after=df_shape_after_batch_op)

    def convert_selected_columns_to_numeric_batch(self):
        self._batch_convert_columns("numeric")

    def convert_selected_columns_to_datetime_batch(self):
        self._batch_convert_columns("datetime")

    def show_context_menu(self, position):
        menu = QMenu()
        set_as_header_action = QAction("Set Selected Row as Header", self)
        set_as_header_action.triggered.connect(self.set_row_as_header)
        menu.addAction(set_as_header_action)
        menu.exec(self.table_view.viewport().mapToGlobal(position))

    @Slot(list, str) # items_info, base_filename
    def prompt_npz_array_selection(self, items_info, base_filename):
        item_display_diag_list = {f"{item['name']} (Shape: {item['shape']}" : f"{item['name']} (Shape: {item['shape']}, Dtype: {item['dtype']})" for item in items_info}
        item_display_list = list(item_display_diag_list.values())
        if not item_display_list:
            if hasattr(self.file_loader_thread, 'worker_instance'):
                self.file_loader_thread.worker_instance.on_npz_array_selected("") # Signal no selection
            return
        input_diag = QInputDialog()
        item_text, ok = input_diag.getItem(self, f"Select Array from {base_filename}",
                                             "Choose an array to load:", list(item_display_diag_list.keys()), 0, False)
        
        item_text = item_display_diag_list.get(item_text, "")
        
        selected_array_name = ""
        if ok and item_text:
            # Extract the original name from the display string (the first part before " (Shape:")
            selected_array_name = item_text.split(" (Shape:")[0]
        
        
        # Signal back to the worker in FileLoaderWorker
        # This assumes AsyncFileLoaderThread has a way to get its worker, or worker has a public method
        if hasattr(self.file_loader_thread, 'worker_instance') and self.file_loader_thread.worker_instance:
             self.file_loader_thread.worker_instance.on_npz_array_selected(selected_array_name)
        elif hasattr(self.file_loader_thread, 'on_npz_array_selected_ relayed'): # if signal is relayed by AsyncFileLoaderThread
             self.file_loader_thread.on_npz_array_selected_relayed.emit(selected_array_name)


    @Slot(list, str, str) # items_info, base_filename, file_type
    def prompt_pickle_item_selection(self, items_info, base_filename, file_type):
        item_display_diag_list = {str(f"{item['name']} (Type: {item['type']}, Shape: {item.get('shape', 'N/A')})")[:60] : f"{item['name']} (Type: {item['type']}, Shape: {item.get('shape', 'N/A')})" for item in items_info}
        item_display_list = list(item_display_diag_list.values())
        if not item_display_list:
            if hasattr(self.file_loader_thread, 'worker_instance'):
                self.file_loader_thread.worker_instance.on_pickle_item_selected("")
            return
        input_diag = QInputDialog()
        input_diag.setFixedSize(400, 300) # Set a fixed size for the dialog
        item_text, ok = input_diag.getItem(self, f"Select Item from {file_type.capitalize()} File: {base_filename}",
                                             "Choose a dataset to load:", list(item_display_diag_list.keys()), 0, False)

        item_text = item_display_diag_list.get(item_text, "")

        selected_item_path = ""
        if ok and item_text:
            # Find the original item's "object_path" based on the selected display string's "name" part
            for item_detail in items_info:
                if item_text.startswith(item_detail['name']):
                    selected_item_path = item_detail['object_path']
                    break
        
        if hasattr(self.file_loader_thread, 'worker_instance') and self.file_loader_thread.worker_instance:
            self.file_loader_thread.worker_instance.on_pickle_item_selected(selected_item_path)
        elif hasattr(self.file_loader_thread, 'on_pickle_item_selected_relayed'):
            self.file_loader_thread.on_pickle_item_selected_relayed.emit(selected_item_path)

    def load_file(self, file_path = None):
        if isinstance(file_path, (str, type(None), bool)) and not file_path:
            self.initial_path = os.path.join(os.path.expanduser('~'), 'Documents')
            if not os.path.exists(self.initial_path): self.initial_path = os.path.join(os.path.expanduser('~'), 'Downloads')
            if not os.path.exists(self.initial_path): self.initial_path = os.path.expanduser('~')
            if hasattr(self, 'initial_path_selected') and self.initial_path_selected and os.path.exists(self.initial_path_selected):
                self.initial_path = self.initial_path_selected
                
            file_dialog_filter = (
                "All Supported Data Files (*.csv *.xlsx *.xls *.xlsb *.xlsm *.parquet *.npz *.pkl *.pickle *.hkl *.hickle *.json *.jsonl *.ndjson *.dat *.txt);;"
                "CSV Files (*.csv);;"
                "Excel Files (*.xlsx *.xls *.xlsb *.xlsm);;"
                "Parquet Files (*.parquet);;"
                "NumPy Archives (*.npz *.npy);;"
                "Pickle Files (*.pkl *.pickle);;"
                "Hickle Files (*.hkl *.hickle);;"
                "JSON Files (*.json *.jsonl *.ndjson);;"
                "Text/DAT Files (*.dat *.txt);;"
                "All Files (*)"
            )
            file_name, _ = QFileDialog.getOpenFileName(self, "Open File", #dir=self.initial_path,
                                                    filter=file_dialog_filter)
            file_name = Path(str(file_name))
            
        else:
            file_name = Path(str(file_path))
                    
        if file_name:
            self.initial_path_selected = os.path.dirname(file_name)
            self._filepath = Path(file_name)
            
            try:
                log_file_dir = os.path.dirname(os.path.abspath(file_name))
                data_basename_no_ext = os.path.splitext(os.path.basename(file_name))[0]
                timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
                log_filename_for_data = f".__log__{data_basename_no_ext}_{timestamp_str}.md"
                self.current_log_file_path = os.path.join(log_file_dir, log_filename_for_data)
                
                logger.set_log_file(self.current_log_file_path, app_name_for_header="Pickaxe", associated_data_file=file_name.name)
                logger.log_action("Pickaxe", "Data Session Start", 
                                  f"Initiated processing for data file: {os.path.basename(file_name)}", 
                                  details={"Log File": self.current_log_file_path})
            except Exception as e:
                print(f"Error setting up logger for {file_name}: {e}")
                # Fallback log if specific one fails
                self.current_log_file_path = os.path.join(os.getcwd(), f".__log__pickaxe_fallback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md")
                logger.set_log_file(self.current_log_file_path, app_name_for_header="Pickaxe (Fallback Log)", associated_data_file=file_name.name)

            sheet_name = None
            if str(file_name).lower().endswith(('.xlsx', '.xlsm', '.xlsb', '.xls')):
                try:
                    sheet_names_list = get_sheet_names(file_name)
                    self.sheet_names = sheet_names_list
                    if len(sheet_names_list) > 1:
                        sheet_name_selected, ok = QInputDialog.getItem(self, "Select Sheet", "Choose a sheet:", sheet_names_list, 0, False)
                        if not ok: self._filepath = Path(''); return
                        sheet_name = sheet_name_selected
                    elif sheet_names_list: sheet_name = sheet_names_list[0]
                    else: self.on_error(f"No sheets found in Excel file: {os.path.basename(file_name)}"); return
                except Exception as e: self.on_error(f"Error reading sheet names: {e}"); return
                
            self.progress_bar.setVisible(True); self.progress_label.setVisible(True)
            self.update_progress(0, f"Loading '{os.path.basename(file_name)}'...")
            QApplication.processEvents()

            self.file_loader_thread = AsyncFileLoaderThread(file_name, sheet_name, parent_gui=self)

            # Connect new signals from the worker instance (which is inside AsyncFileLoaderThread)
            # This assumes AsyncFileLoaderThread exposes its worker or the signals directly.
            # Let's assume AsyncFileLoaderThread's worker is self.worker_instance
            if hasattr(self.file_loader_thread, 'worker_instance'): # If worker is accessible
                self.file_loader_thread.worker_instance.request_npz_array_selection.connect(self.prompt_npz_array_selection)
                self.file_loader_thread.worker_instance.request_pickle_item_selection.connect(self.prompt_pickle_item_selection)
                self.file_loader_thread.worker_instance.request_excel_to_csv_conversion.connect(self.prompt_excel_to_csv_conversion)
            else: # If AsyncFileLoaderThread itself relays the signals
                self.file_loader_thread.request_npz_array_selection.connect(self.prompt_npz_array_selection)
                self.file_loader_thread.request_pickle_item_selection.connect(self.prompt_pickle_item_selection)
                self.file_loader_thread.request_excel_to_csv_conversion.connect(self.prompt_excel_to_csv_conversion) 

            self.file_loader_thread.finished_signal.connect(self.on_file_loaded)
            self.file_loader_thread.error_signal.connect(self.on_error)
            self.file_loader_thread.start()

        else:
            self.update_progress(0, ""); self.progress_bar.setVisible(False); self.progress_label.setVisible(False)

    @Slot(str)
    def prompt_excel_to_csv_conversion(self, filename):
        reply = QMessageBox.question(self, "Excel Load Failed",
                                       f"Could not open '{os.path.basename(filename)}' directly.\n\n"
                                       "This can happen with very large or complex files.\n\n"
                                       "Do you want to try converting it to a temporary CSV file and loading that instead?",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        
        choice = (reply == QMessageBox.StandardButton.Yes)
        
        if hasattr(self.file_loader_thread, 'on_excel_conversion_choice_relayed'):
            self.file_loader_thread.on_excel_conversion_choice_relayed(choice)

    @Slot(object, dict)
    def on_file_loaded(self, df, file_info):
        
        self.types_suggested_and_applied_this_session = False
        
        if not self.current_log_file_path: # Should have been set in load_file
             # Fallback if somehow not set (e.g. direct call to on_file_loaded without load_file)
            fallback_log_name = f".__log__{os.path.splitext(os.path.basename(self._filepath))[0] if self._filepath else 'unknown_data'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
            self.current_log_file_path = os.path.join(os.getcwd(), fallback_log_name)
            logger.set_log_file(self.current_log_file_path, "Pickaxe", associated_data_file=self._filepath.name or "Unknown")

        if df is None or df.is_empty():
            
            logger.log_action(
                "Pickaxe", "File Load Error", 
                f"Failed to load or file empty: {os.path.basename(self._filepath if self._filepath else 'Unknown')}",
                details=file_info 
            )
            self.df_shape_before = (0, 0)
            self.on_error(f"File '{os.path.basename(self._filepath)}' loaded empty or failed.")
            self.df = None; self.filtered_df = None
            self.update_model_slot(None, True); self.update_file_info_label()
            self.save_button.setEnabled(False); self.apply_query_button.setEnabled(False)
            self.filter_toggle_button.setEnabled(False); self.reset_filters_button.setEnabled(False)
            self.prev_columns_button.setEnabled(False); self.next_columns_button.setEnabled(False)
            if self.filter_panel: self.filter_panel.setEnabled(False)
            self.update_completer_model()
            return
        
        if "__original_index__" not in df.columns:
            self.df = df.with_row_count("__original_index__") # Add original index
        
        if self.df is not None:
            self.filtered_df = self.df.clone()

        if self._filepath.name.lower().endswith(('.xlsx', '.xlsm', '.xlsb', '.xls')):
            file_info['total_sheets'] = len(self.sheet_names) if hasattr(self, 'sheet_names') else (1 if file_info.get('sheet_name') else 'N/A')
        self.file_info = file_info
        if self.df is not None:
            self.original_row_count = self.df.height 
        
        logger.log_dataframe_load(
            "Pickaxe",
            self.file_info.get('filename', self._filepath), 
            sheet_name=self.file_info.get('sheet_name'),
            rows=self.file_info.get('rows', df.height), 
            cols=self.file_info.get('columns', df.width), 
            load_time_sec=self.file_info.get('load_time', 0)
        )
        if self.df is not None:
            self.df_shape_before = (self.df.height, self.df.width)
                            
        self.applied_filters_info = []
        self.update_model_slot(self.filtered_df, True)
        self.update_file_info_label()
        if self.filter_panel:
            self.filter_panel.panel_filters_changed.disconnect(self.update_query_input_from_structured_filters)
            self.filter_panel.deleteLater(); self.filter_panel = None

        if self.df is not None:
            columns_for_panel = [c for c in self.df.columns if c != "__original_index__"]
            self.filter_panel = FilterPanel(columns_for_panel)
            self.filter_panel.panel_filters_changed.connect(self.update_query_input_from_structured_filters)

            if columns_for_panel or not df.is_empty(): # Always enable filter panel if df is loaded
                self.filter_panel.apply_button.clicked.connect(self.apply_structured_filters)
                self.filter_panel.setEnabled(True); self.filter_toggle_button.setEnabled(True)
            else: 
                self.filter_panel.setEnabled(False)
                self.filter_toggle_button.setEnabled(False)
                
            self._layout.insertWidget(4, self.filter_panel)
            self.filter_panel.setVisible(False)
            
        self.update_progress(100, f"File '{os.path.basename(self._filepath)}' loaded.")
        if hasattr(self, 'file_loader_thread') and self.file_loader_thread:
            self.file_loader_thread.quit(); self.file_loader_thread.wait()

        if df is not None and not df.is_empty():
            self.file_button.setDefault(False)

            self.save_button.setEnabled(True)
            self.save_button.setDefault(True)

            self.apply_query_button.setEnabled(True)
            self.filter_toggle_button.setEnabled(True)
            self.reset_filters_button.setEnabled(True)
            self.vw_button.setEnabled(True)
            self.dt_button.setEnabled(True)
            self.suggest_conversions_button.setEnabled(True) 

        else:
            self.file_button.setDefault(True)

            self.save_button.setEnabled(False)
            self.save_button.setDefault(False)

            self.apply_query_button.setEnabled(False)
            self.filter_toggle_button.setEnabled(False)
            self.reset_filters_button.setEnabled(False)
            self.vw_button.setEnabled(False) # Disable VW button if no data
            self.dt_button.setEnabled(False)
            self.suggest_conversions_button.setEnabled(False) # Disable new button
            
        if self.df is not None:
            displayable_width = len([c for c in self.df.columns if c != "__original_index__"])
            self.prev_columns_button.setEnabled(displayable_width > self.columns_per_page)
            self.next_columns_button.setEnabled(displayable_width > self.columns_per_page)

        self.update_completer_model()

        if self.table_view.selectionModel().hasSelection():
            current_selection = self.table_view.selectionModel().currentIndex()
            if current_selection.isValid(): self.handle_column_selection_for_stats(current_selection, None)
        else: self.stats_label.setText("Select a column to see statistics."); self.filled_rows_bar.set_fill_percentage(0); self.set_stats_label_font_monospaced()


    def suggest_and_convert_types(self):
        df_to_analyze = self._get_df_to_operate_on()
        if df_to_analyze is None:
            QMessageBox.information(self, "No Data", "No data loaded to analyze for conversions.")
            return

        self.update_progress(0, "Analyzing columns for potential type conversions...")
        QApplication.processEvents()

        suggested_conversions = [] 
        MAX_UNIQUE_CATEGORICAL_ABS = 50 
        MAX_UNIQUE_CATEGORICAL_REL = 0.1 
        BOOLEAN_TRUE_STRINGS = {"True", "true", "yes", "Yes"}
        BOOLEAN_FALSE_STRINGS = {"False", "false", "no", "No"}


        for col_idx, col_name in enumerate([l for l in df_to_analyze.columns if l != "__original_index__"]):
            self.update_progress(int(((col_idx +1) / (df_to_analyze.width -1 or 1)) * 100) , f"Analyzing: {col_name}")
            QApplication.processEvents()
            
            series = df_to_analyze.get_column(col_name)
            non_null_series = series.drop_nulls()
            
            current_suggestion = "string_key" 

            if non_null_series.is_empty():
                suggested_conversions.append((col_name, current_suggestion)) 
                continue

            # 1. Check existing non-string types
            if series.dtype in [pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64]:
                current_suggestion = "integer"
            elif series.dtype in [pl.Float32, pl.Float64]:
                current_suggestion = "numeric_float"
            elif series.dtype == pl.Categorical:
                current_suggestion = "category"
            elif series.dtype in [pl.Date, pl.Datetime]:
                current_suggestion = "datetime"
            elif series.dtype == pl.Boolean:
                current_suggestion = "boolean" # Already boolean
            
            # 2. If it's a string, try to infer better types
            elif series.dtype == pl.Utf8:
                # Attempt Datetime
                try:
                    sample_size_dt = min(1000, non_null_series.len())
                    casted_dt_sample = non_null_series.sample(sample_size_dt).str.to_datetime(strict=False, time_unit='us', ambiguous='earliest', format=None)
                    if casted_dt_sample.null_count() / sample_size_dt <= 0.1:
                         casted_dt_full = non_null_series.str.to_datetime(strict=False, time_unit='us', ambiguous='earliest', format=None)
                         if casted_dt_full.null_count() / non_null_series.len() <= 0.05: 
                             current_suggestion = "datetime"
                except Exception: pass

                # Attempt Boolean if not datetime
                if current_suggestion == "string_key":
                    unique_lower_strings = {s.lower() for s in non_null_series.unique().to_list() if s is not None}
                    if unique_lower_strings.issubset(BOOLEAN_TRUE_STRINGS.union(BOOLEAN_FALSE_STRINGS)):
                        current_suggestion = "boolean"

                # Attempt Integer if not datetime or boolean
                if current_suggestion == "string_key":
                    try:
                        _ = non_null_series.cast(pl.Int64, strict=True) 
                        current_suggestion = "integer"
                    except: 
                        try:
                            casted_float_for_int_check = non_null_series.cast(pl.Float64, strict=False)
                            if not casted_float_for_int_check.is_null().all() and \
                               (casted_float_for_int_check.drop_nulls().fill_null(0.1) % 1 == 0).all(): # check whole numbers
                                if casted_float_for_int_check.null_count() / non_null_series.len() <= 0.05:
                                    current_suggestion = "integer"
                            # Attempt Float (lenient) if not integer-like float
                            elif current_suggestion == "string_key": 
                                if casted_float_for_int_check.null_count() / non_null_series.len() <= 0.05:
                                    current_suggestion = "numeric_float"
                        except Exception: pass

                # Attempt Categorical if not datetime, boolean, or numeric
                if current_suggestion == "string_key": 
                    n_unique = non_null_series.n_unique()
                    if n_unique <= MAX_UNIQUE_CATEGORICAL_ABS or \
                       (non_null_series.len() > 0 and (n_unique / non_null_series.len()) <= MAX_UNIQUE_CATEGORICAL_REL):
                        current_suggestion = "category"
            
            suggested_conversions.append((col_name, current_suggestion))

        self.update_progress(100, "Presenting suggestions...")
        QApplication.processEvents()

        if not suggested_conversions:
            self.update_progress(0, "No columns found to analyze or suggest conversions for.")
            QMessageBox.information(self, "No Suggestions", "No columns available for conversion suggestions.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Suggested Type Conversions")
        dialog_layout = QVBoxLayout(dialog)

        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel("<b>Column Name</b>"), 2)
        # New Order: String/Key, Category, Boolean, Integer, Numeric (Float), Datetime
        header_layout.addWidget(QLabel("<b>String/Key</b>"), 1, alignment=Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(QLabel("<b>Category</b>"), 1, alignment=Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(QLabel("<b>Boolean</b>"), 1, alignment=Qt.AlignmentFlag.AlignCenter) # New
        header_layout.addWidget(QLabel("<b>Integer</b>"), 1, alignment=Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(QLabel("<b>Numeric (Float)</b>"), 1, alignment=Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(QLabel("<b>Datetime</b>"), 1, alignment=Qt.AlignmentFlag.AlignCenter)
        dialog_layout.addLayout(header_layout)

        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        grid_layout = QGridLayout(scroll_widget)
        grid_layout.setColumnStretch(0, 2) 
        for i in range(1, 7): grid_layout.setColumnStretch(i, 1) # Now 6 type columns


        self.conversion_radio_button_groups = [] 

        for row_idx, (col_name, suggested_type) in enumerate(suggested_conversions):
            col_name_label = QLabel(col_name)
            string_rb = QRadioButton(); category_rb = QRadioButton(); boolean_rb = QRadioButton() # New
            integer_rb = QRadioButton(); float_rb = QRadioButton(); datetime_rb = QRadioButton()


            button_group = QButtonGroup(dialog)
            button_group.addButton(string_rb); button_group.addButton(category_rb); button_group.addButton(boolean_rb)
            button_group.addButton(integer_rb); button_group.addButton(float_rb); button_group.addButton(datetime_rb)

            if suggested_type == "string_key": string_rb.setChecked(True)
            elif suggested_type == "category": category_rb.setChecked(True)
            elif suggested_type == "boolean": boolean_rb.setChecked(True) # New
            elif suggested_type == "integer": integer_rb.setChecked(True)
            elif suggested_type == "numeric_float": float_rb.setChecked(True)
            elif suggested_type == "datetime": datetime_rb.setChecked(True)
            else: string_rb.setChecked(True) 

            grid_layout.addWidget(col_name_label, row_idx, 0)
            grid_layout.addWidget(string_rb, row_idx, 1, alignment=Qt.AlignmentFlag.AlignCenter)
            grid_layout.addWidget(category_rb, row_idx, 2, alignment=Qt.AlignmentFlag.AlignCenter)
            grid_layout.addWidget(boolean_rb, row_idx, 3, alignment=Qt.AlignmentFlag.AlignCenter) # New
            grid_layout.addWidget(integer_rb, row_idx, 4, alignment=Qt.AlignmentFlag.AlignCenter)
            grid_layout.addWidget(float_rb, row_idx, 5, alignment=Qt.AlignmentFlag.AlignCenter)
            grid_layout.addWidget(datetime_rb, row_idx, 6, alignment=Qt.AlignmentFlag.AlignCenter)


            self.conversion_radio_button_groups.append(
                (col_name_label, string_rb, category_rb, boolean_rb, integer_rb, float_rb, datetime_rb)
            )
        
        scroll_widget.setLayout(grid_layout)
        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        dialog_layout.addWidget(scroll_area)

        button_box_layout = QHBoxLayout() 
        help_button = QPushButton("?")
        help_button.setFixedSize(25, 25)
        help_button.setToolTip("Help on Data Types")
        help_button.clicked.connect(self._show_type_conversion_help) 
        
        dialog_buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        dialog_buttons.accepted.connect(dialog.accept)
        dialog_buttons.rejected.connect(dialog.reject)

        button_box_layout.addWidget(help_button)
        button_box_layout.addStretch() 
        button_box_layout.addWidget(dialog_buttons)
        dialog_layout.addLayout(button_box_layout) 

        dialog.setMinimumWidth(750) # Wider for the new column
        dialog.setMinimumHeight(400)

        dialog.accepted.connect(self._handle_type_suggestion_accepted)
        dialog.rejected.connect(self._handle_type_suggestion_rejected)
        dialog.show()

    def _handle_type_suggestion_accepted(self):
        cols_to_float = []; cols_to_int = []; cols_to_datetime = []
        cols_to_category = []; cols_to_boolean = []; cols_to_string = []

        for col_label, str_rb, cat_rb, bool_rb, int_rb, flt_rb, dt_rb in self.conversion_radio_button_groups:
            col_name_to_convert = col_label.text()
            
            if self.df is not None:
                current_col_dtype = self.df.schema[col_name_to_convert]
                if flt_rb.isChecked() and not isinstance(current_col_dtype, (pl.Float32, pl.Float64)):
                    cols_to_float.append(col_name_to_convert)
                elif int_rb.isChecked() and not current_col_dtype.is_integer():
                    cols_to_int.append(col_name_to_convert)
                elif dt_rb.isChecked() and not current_col_dtype.is_temporal():
                    cols_to_datetime.append(col_name_to_convert)
                elif cat_rb.isChecked() and current_col_dtype != pl.Categorical:
                    cols_to_category.append(col_name_to_convert)
                elif bool_rb.isChecked() and current_col_dtype != pl.Boolean: # New
                    cols_to_boolean.append(col_name_to_convert)
                elif str_rb.isChecked() and current_col_dtype != pl.Utf8: 
                    cols_to_string.append(col_name_to_convert)
            
        any_conversion_done = False
        if cols_to_float:
            self.context_menu_selected_column_names = cols_to_float
            self._batch_convert_columns("numeric_float", target_polars_type=pl.Float64)
            any_conversion_done = True
        if cols_to_int:
            self.context_menu_selected_column_names = cols_to_int
            self._batch_convert_columns("integer", target_polars_type=pl.Int64)
            any_conversion_done = True
        if cols_to_datetime:
            self.context_menu_selected_column_names = cols_to_datetime
            self._batch_convert_columns("datetime", target_polars_type=pl.Datetime)
            any_conversion_done = True
        if cols_to_category:
            self.context_menu_selected_column_names = cols_to_category
            self._batch_convert_columns("categorical", target_polars_type=pl.Categorical)
            any_conversion_done = True
        if cols_to_boolean: # New
            self.context_menu_selected_column_names = cols_to_boolean
            self._batch_convert_columns("boolean", target_polars_type=pl.Boolean)
            any_conversion_done = True
        if cols_to_string: 
            self.context_menu_selected_column_names = cols_to_string
            self._batch_convert_columns("string_key", target_polars_type=pl.Utf8)
            any_conversion_done = True
        
        if any_conversion_done:
            self.types_suggested_and_applied_this_session = True
            self.update_model_slot(self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df, False) 
            self.update_completer_model() 
            self.update_progress(100, "Selected type conversions applied.")
        else:
            self.update_progress(100, "No changes made to data types.")

        if self.visual_workshop_window is not None and self.visual_workshop_window.isVisible():
            self.continue_opening_visual_workshop()
            
        
    def _handle_type_suggestion_rejected(self):
            self.update_progress(100, "Conversion suggestions cancelled.")
            # If user cancels suggest_types, we might still proceed or not
            if not self.types_suggested_and_applied_this_session: # Check if they actually applied changes
                if QMessageBox.question(self, "Proceed?", "Proceed to Visual Workshop without applying type suggestions?",
                                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No) == QMessageBox.StandardButton.No:
                    self.statusBar().showMessage("Visual Workshop launch cancelled.", 2000)
                else:
                    if self.visual_workshop_window is not None and self.visual_workshop_window.isVisible():
                        self.statusBar().showMessage("Type suggestions ignored, proceeding to Visual Workshop.", 2000)
                        self.continue_opening_visual_workshop()

    def _show_type_conversion_help(self):
        help_text = """
        <html>
        <head>
            <style>
                body { font-family: sans-serif; font-size: 10pt; }
                h3 { color: #333; margin-top: 10px; margin-bottom: 3px;}
                p { margin-bottom: 8px; }
                ul { margin-top: 0px; padding-left: 20px; }
                li { margin-bottom: 4px; }
                code { background-color: #f0f0f0; padding: 1px 3px; border-radius: 3px; font-family: monospace;}
            </style>
        </head>
        <body>
            <h3>Understanding Data Types for Conversion:</h3>
            
            <p><b>String/Key:</b>
                <ul>
                    <li><b>What it is:</b> Text data. Can be words, sentences, codes, or numbers not used for math (like IDs, phone numbers).</li>
                    <li><b>When to use:</b> For unique identifiers, free-form text, or when other types don't fit. If a column has mixed letters and numbers, or numbers you'll never do math on, String/Key is often best.</li>
                    <li><b>Example:</b> "ABC-123", "John Doe", "This is a note", "90210" (as ZIP), "POL_10023".</li>
                </ul>
            </p>

            <p><b>Category:</b>
                <ul>
                    <li><b>What it is:</b> Text or numbers representing a fixed, limited set of groups or labels. Useful for columns with few unique values that repeat often.</li>
                    <li><b>When to use:</b> When values define distinct groups (e.g., "Gender", "Country Code", "Product Type", "Status"). This can save memory and speed up operations like grouping or filtering. Great for coloring or faceting in plots.</li>
                    <li><b>Example:</b> "Male"/"Female", "USA"/"CAN"/"MEX", "Active"/"Inactive", "Yes"/"No", "Type A"/"Type B".</li>
                </ul>
            </p>

            <p><b>Boolean:</b>
                <ul>
                    <li><b>What it is:</b> Represents true or false states.</li>
                    <li><b>When to use:</b> For columns that only have two states, often represented as True/False, Yes/No, 1/0, T/F. Converting to a true Boolean type is efficient and clear.</li>
                    <li><b>Example:</b> Values like <code>true</code>, <code>false</code>, <code>0</code>, <code>1</code>, <code>YES</code>, <code>NO</code>.</li>
                </ul>
            </p>

            <p><b>Integer (Whole Number):</b>
                <ul>
                    <li><b>What it is:</b> Whole numbers without decimal points (e.g., -2, 0, 5, 100).</li>
                    <li><b>When to use:</b> For counts, quantities, or discrete numerical values you might perform math on. If a column only contains numbers like 1, 2, 150, -10 (and no 1.5, 2.3), Integer is suitable.</li>
                    <li><b>Example:</b> 42, -100, 0, 78932, number of items.</li>
                </ul>
            </p>

            <p><b>Numeric (Float):</b>
                <ul>
                    <li><b>What it is:</b> Numbers that can have decimal points. Used for continuous measurements, amounts, calculations.</li>
                    <li><b>When to use:</b> For any number you might do math with, especially if it can have fractions or requires high precision (e.g., prices, percentages, sensor readings, monetary values).</li>
                    <li><b>Example:</b> 123.45, -0.5, 100.00, 3.14159.</li>
                </ul>
            </p>

            <p><b>Datetime:</b>
                <ul>
                    <li><b>What it is:</b> Represents specific dates, and optionally, times.</li>
                    <li><b>When to use:</b> For any column that holds date or date & time information. This allows for date-based calculations (e.g., time differences, extracting year/month) and proper plotting on time axes.</li>
                    <li><b>Example:</b> "2023-10-26", "2023-10-26 14:30:00", "10/26/2023". The tool will try to understand common date formats.</li>
                </ul>
            </p>
            <p><i>Choosing the right type helps with data analysis, plotting, and memory usage. The suggestions are based on the data's appearance, but your knowledge of the data is key! Converting to a more specific type can enable more powerful operations.</i></p>
        </body>
        </html>
        """
        
        help_dialog = QMessageBox(self)
        help_dialog.setWindowTitle("Data Type Conversion Help")
        help_dialog.setTextFormat(Qt.TextFormat.RichText) 
        help_dialog.setText(help_text)
        help_dialog.setIcon(QMessageBox.Icon.Information)
        help_dialog.setStandardButtons(QMessageBox.StandardButton.Ok)
        # Make dialog wider to better display help text
        help_dialog.setStyleSheet("QTextEdit{ min-width: 450px; min-height: 300px}")
        help_dialog.exec()


    @Slot(object, bool)
    def update_model_slot(self, df_to_display, reset_col_page=False):
        if df_to_display is None or df_to_display.is_empty():
            self.table_view.setModel(None); self.model = None
            self.stats_label.setText("No data to display."); self.filled_rows_bar.set_fill_percentage(0)
            self.set_stats_label_font_monospaced(); self.update_column_range_label(None)
            self.prev_columns_button.setEnabled(False); self.next_columns_button.setEnabled(False)
            return

        if reset_col_page: self.current_column_page = 0

        all_displayable_columns = [c for c in df_to_display.columns if c != "__original_index__"]
        start_display_col_idx = self.current_column_page * self.columns_per_page
        end_display_col_idx = min(start_display_col_idx + self.columns_per_page, len(all_displayable_columns))
        current_page_display_columns = all_displayable_columns[start_display_col_idx:end_display_col_idx]

        cols_for_model_subset = current_page_display_columns[:]
        if "__original_index__" in df_to_display.columns:
            cols_for_model_subset.append("__original_index__")
        else: # Should not happen if loaded correctly, but as a safeguard
            print("Warning: __original_index__ missing in df_to_display for model update")


        row_limit_print = min(1000, df_to_display.height)
        # Select only the necessary columns for the current page + index, then slice rows
        subset_df_for_model = df_to_display.select(cols_for_model_subset).slice(0, row_limit_print)

        self.model = PolarsModel(subset_df_for_model)
        self.table_view.setModel(self.model)

        if self.table_view.selectionModel():
            # try: self.table_view.selectionModel().currentColumnChanged.disconnect(self.handle_column_selection_for_stats)
            # except RuntimeError: pass
            self.table_view.selectionModel().currentColumnChanged.connect(self.handle_column_selection_for_stats)

        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table_view.verticalHeader().setDefaultSectionSize(8)
        self.update_column_range_label(df_to_display) # Pass the full df_to_display for accurate total counts

        self.prev_columns_button.setEnabled(self.current_column_page > 0)
        self.next_columns_button.setEnabled(end_display_col_idx < len(all_displayable_columns))


    def show_previous_columns(self):
        df_for_nav = self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df
        if df_for_nav is None: return
        if self.current_column_page > 0:
            self.current_column_page -= 1
            self.progress_bar.setVisible(True); self.progress_label.setVisible(True)
            self.update_progress(0, "Loading previous columns...")
            QApplication.processEvents(); self.threaded_update_model()
            self.update_progress(100, "Previous columns loaded.")

    def show_next_columns(self):
        df_for_nav = self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df
        if df_for_nav is None: return

        all_displayable_columns = [c for c in df_for_nav.columns if c != "__original_index__"]
        if (self.current_column_page + 1) * self.columns_per_page < len(all_displayable_columns):
            self.current_column_page += 1
            self.progress_bar.setVisible(True); self.progress_label.setVisible(True)
            self.update_progress(0, "Loading next columns...")
            QApplication.processEvents(); self.threaded_update_model()
            self.update_progress(100, "Next columns loaded.")


    def threaded_update_model(self):
        df_to_use = self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df
        self.update_model_signal.emit(df_to_use, False)

    @Slot(int, str)
    def update_progress(self, value, message):
        self.progress_label.setText(message)
        self.progress_bar.setValue(value)
        if value == 100:
            QApplication.processEvents()
            QTimer.singleShot(250, lambda: (self.progress_bar.setVisible(False), self.progress_label.setVisible(False)))
        else: self.progress_bar.setVisible(True); self.progress_label.setVisible(True)

    @Slot(str)
    def on_error(self, error_message):
        if hasattr(self, 'file_loader_thread') and self.file_loader_thread and self.file_loader_thread.isRunning():
            self.file_loader_thread.quit(); self.file_loader_thread.wait()
        self.progress_bar.setVisible(False); self.progress_label.setVisible(False)
        QMessageBox.critical(self, "Error", error_message)

    def load_dataframe_from_source(self, df_from_source, name_hint, source_log_file_path=None):
        self.statusBar().showMessage(f"Receiving data: {name_hint}", 3000)
        self._filepath = Path(name_hint) 
        self.df = df_from_source.clone()
        if "__original_index__" not in self.df.columns:
            self.df = self.df.with_row_count("__original_index__")
        self.filtered_df = self.df.clone()
        
        if source_log_file_path and os.path.exists(os.path.dirname(source_log_file_path)):
            self.current_log_file_path = source_log_file_path
            logger.set_log_file(self.current_log_file_path, "Pickaxe (Continuing Session)", associated_data_file=name_hint)
        else: # Fallback if no valid log path from source
            log_file_dir = os.getcwd()
            data_basename_no_ext = re.sub(r'\W+', '_', name_hint) 
            timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_filename_for_data = f".__log__{data_basename_no_ext}_{timestamp_str}.md"
            self.current_log_file_path = os.path.join(log_file_dir, log_filename_for_data)
            logger.set_log_file(self.current_log_file_path, "Pickaxe (New Session for Transformed Data)", associated_data_file=name_hint)

        logger.log_action("Pickaxe", "Data Received", f"Data '{name_hint}' received into Pickaxe.",
                          details={"Source": "Data Transformer", "Shape": self.df.shape, "Log File": self.current_log_file_path})

        # Update UI (similar to end of on_file_loaded)
        simulated_file_info = {
            'filename': name_hint, 'rows': self.df.height, 
            'columns': self.df.width - (1 if "__original_index__" in self.df.columns else 0),
            'sheet_name': "From Data Transformer"
        }
        self.file_info = simulated_file_info
        self.original_row_count = self.df.height
        self.applied_filters_info = []
        self.query_input.clear()
        self.update_model_slot(self.filtered_df, True)
        self.update_file_info_label()
        self.update_completer_model()
        # ... (reinitialize filter panel, enable buttons as in on_file_loaded) ...
        self.update_progress(100, f"Data '{name_hint}' loaded from Data Transformer.")


    def save_file(self): # This is the original button-triggered save
        if self.df is None: QMessageBox.warning(self, "Save Error", "No data loaded to save."); return
        original_name = os.path.splitext(os.path.basename(self._filepath if self._filepath else "data"))[0]
        # ... (filter_suffix and suggested_name logic - duplicated from above, consider refactoring)
        filter_info_str_parts = []
        for item in self.applied_filters_info:
            if isinstance(item, tuple):
                column, comparison, fs1, fs2, _, ftype, regex, negate = item
                fs1_clean = re.sub(r'[^a-zA-Z0-9_-]', '', str(fs1))[:10]
                fs2_clean = re.sub(r'[^a-zA-Z0-9_-]', '', str(fs2))[:10] if fs2 else ""
                comp_short = comparison.replace(" ", "")[:10].lower() 
                col_str = str(column)[:6] if column else "Row"
                part = f"{col_str}_{'N' if negate else ''}{'R' if regex else ''}{ftype[:1]}_{comp_short}_{fs1_clean}"
                if fs2_clean: part += f"_{fs2_clean}"
                filter_info_str_parts.append(part)
            elif isinstance(item, str) and item.startswith("Query: "):
                query_clean = re.sub(r'[^a-zA-Z0-9_=-]', '', item.replace("Query: ", ""))[:20]
                filter_info_str_parts.append(f"Q_{query_clean}")
            elif isinstance(item, str): filter_info_str_parts.append(item)
        filter_suffix = "_".join(filter_info_str_parts)
        filter_suffix = re.sub(r'_+', '_', filter_suffix).strip('_')
        suggested_name = f"{original_name}_processed.csv" if not filter_suffix else f"{original_name}_filtered_{filter_suffix}.csv"
        max_len_suggested = 200
        if len(suggested_name) > max_len_suggested:
            ext = os.path.splitext(suggested_name)[1]
            suggested_name = suggested_name[:max_len_suggested - len(ext)] + ext
        
        file_name_to_save, selected_filter = QFileDialog.getSaveFileName(self, "Save File", suggested_name, "CSV Files (*.csv);;Excel Files (*.xlsx)")
        
        if file_name_to_save:
            if self.current_log_file_path:
                 logger.set_log_file(self.current_log_file_path, "Pickaxe") # Ensure correct log context
            logger.log_action("Pickaxe", "File Save Initiated", f"User initiated save to '{os.path.basename(file_name_to_save)}'.",
                details={"Path": file_name_to_save, "Format": selected_filter, 
                         "Source Data": os.path.basename(self._filepath) if self._filepath else "Current Data"})

            self.progress_bar.setVisible(True); self.progress_label.setVisible(True); self.progress_bar.setValue(0)
            self.df_to_save_op_ref = self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df
            if self.df_to_save_op_ref is None:
                QMessageBox.warning(self, "Save Error", "No data available to save.")
                self.progress_bar.setVisible(False); self.progress_label.setVisible(False); return

            self.file_saver_thread = FileSaverWorker(self.df_to_save_op_ref.clone(), file_name_to_save, self.update_progress, self.on_save_completed, self.on_save_error, self.file_info)
            self.file_saver_thread.start()

    @Slot(str)
    def on_save_completed(self, file_name_cb):
        
        if self.df is not None:
            logger.log_dataframe_save(
                "Pickaxe", file_name_cb,
                rows=getattr(self, 'df_to_save_op_ref', self.df).height, 
                cols=getattr(self, 'df_to_save_op_ref', self.df).width,
                source_data_name=os.path.basename(self._filepath) if self._filepath else "current data"
            )
        
        self.update_progress(100, f"File saved successfully as {file_name_cb}")
        QMessageBox.information(self, "Save Successful", f"File saved successfully as {file_name_cb}")
        if hasattr(self, 'file_saver_thread'): self.file_saver_thread.quit(); self.file_saver_thread.wait()

    @Slot(str)
    def on_save_error(self, error_message):
        self.update_progress(100, f"Error saving file: {error_message}")
        QMessageBox.critical(self, "Save Error", f"Error saving file: {error_message}")
        if hasattr(self, 'file_saver_thread'): self.file_saver_thread.quit(); self.file_saver_thread.wait()

    def apply_structured_filters(self):
        if self.df is not None and self.filter_panel is not None:
            filters_and_logic = self.filter_panel.get_filters_and_logic()
            if not filters_and_logic or not any(isinstance(item, tuple) for item in filters_and_logic):
                self.reset_all_filters() # If no filters defined, reset to original
                self.update_progress(100, "No filters applied. Showing original data.")
                return

            self.applied_filters_info = filters_and_logic
            self.update_query_input_from_structured_filters()

            self.progress_bar.setVisible(True); self.progress_label.setVisible(True); self.progress_bar.setValue(0)

            df_shape_before_filter = self.df.shape # df is the basis for filtering
            # Store this to pass to on_filter_completed, or log here and then in on_filter_completed for the result
            self.df_shape_before_current_op = df_shape_before_filter 
            
            # Pass self.df (which has __original_index__) to FilterWorker
            self.filter_worker = FilterWorker(self.df.clone(), filters_and_logic)
            self.filter_worker.progress_updated.connect(self.update_progress)
            self.filter_worker.filter_completed.connect(self.on_filter_completed)
            self.filter_worker.start()

    def apply_polars_query(self):
        if self.df is None: QMessageBox.information(self, "Query Error", "No data loaded to query."); return
        
        df_shape_before_query = self.df.shape # Query is applied on self.df
        query_string = self.query_input.text().strip()
        
        self.df_shape_before = self.df.shape
        
        # Use self.df (which includes __original_index__) for context,
        # but the user's query should ideally not reference __original_index__ directly.
        # The query will operate on the DataFrame, and __original_index__ will be preserved.
        eval_context = {"pl": pl, "datetime": datetime, "date": date}
        if self.df is not None:
            # Provide displayable columns in context for pl.col() usage
            display_cols = [c for c in self.df.columns if c != "__original_index__"]
            eval_context.update({col: pl.col(col) for col in display_cols})


        if not query_string:
            self.reset_all_filters()
            self.update_progress(100, "Query cleared. Showing original data.")
            return

        self.progress_bar.setVisible(True); self.progress_label.setVisible(True)
        self.update_progress(0, f"Applying query: {query_string[:50]}...")
        QApplication.processEvents()
        try:
            # The filter expression is applied to self.df
            filter_expression = eval(query_string, eval_context)
            self.filtered_df = self.df.filter(filter_expression)
            self.applied_filters_info = [f"Query: {query_string}"]
            self.update_model_slot(self.filtered_df, True)
            self.update_progress(100, "Query applied successfully.")
            if self.table_view.selectionModel().hasSelection():
                current_selection = self.table_view.selectionModel().currentIndex()
                if current_selection.isValid(): self.handle_column_selection_for_stats(current_selection, None)
                
            logger.log_action(
                "Pickaxe", "Polars Query Applied", f"Query: {query_string}",
                details={"Full Query": query_string},
                df_shape_before=df_shape_before_query, 
                df_shape_after=self.filtered_df.shape if self.filtered_df is not None else None
            )

        except Exception as e:
            self.update_progress(100, "Query failed.")
            QMessageBox.critical(self, "Polars Query Error", f"Error executing query:\n'{query_string}'\n\n{str(e)}")
            logger.log_action("Pickaxe", "Polars Query Error",  f"Query: {query_string}", details={"Error": str(e), "Full Query": query_string})
        

    def show_query_help_extended(self):
        # Create a small dialog or message box offering multiple links
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Polars Query Help")
        msg_box.setTextFormat(Qt.TextFormat.RichText) # Allow HTML for links
        msg_box.setText(
            "<b>Polars Expressions:</b><br>"
            "<a href='https://docs.pola.rs/user-guide/expressions/'>Official Polars Expression Guide</a><br><br>"
            "<b>Regex (Regular Expressions):</b><br>"
            "Used in Polars string functions like <code>.str.contains(r'pattern')</code>, <code>.str.extract()</code>, etc.<br>"
            "Learn Regex at: <a href='https://www.regexone.com/'>regexone.com</a>"
        )
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.exec()

    def update_column_range_label(self, df_with_potential_index):
        if df_with_potential_index is not None and not df_with_potential_index.is_empty():
            displayable_columns = [c for c in df_with_potential_index.columns if c != "__original_index__"]
            total_display_cols_active = len(displayable_columns)

            start_display_col_idx = self.current_column_page * self.columns_per_page
            # end_display_col_idx is exclusive, so it's the number of cols shown up to this point.
            end_display_col_idx_on_page = min(start_display_col_idx + self.columns_per_page, total_display_cols_active)

            start_col_display_num = start_display_col_idx + 1 if total_display_cols_active > 0 else 0
            end_col_display_num = end_display_col_idx_on_page if total_display_cols_active > 0 else 0

            current_rows_active = df_with_potential_index.height # Total rows in current (filtered) view

            row_info = f"   Rows: {current_rows_active:,}"
            # self.original_row_count is from the initially loaded self.df (before __original_index__)
            # For consistency, use self.df.height (which has __original_index__) if available
            original_total_rows = self.df.height if self.df is not None else 0

            if original_total_rows > 0 and current_rows_active != original_total_rows:
                percentage = (current_rows_active / original_total_rows) * 100
                percentage_str = f"{percentage:.1f}"
                if percentage == 100.0: percentage_str = "100"
                elif 0 < percentage < 0.1: percentage_str = f"{percentage:.3f}"
                elif percentage == 0.0: percentage_str = "0"
                row_info = f"   Rows: {current_rows_active:,} ({percentage_str}% of {original_total_rows:,})"

            col_info = f"Display Columns {start_col_display_num}-{end_col_display_num} of {total_display_cols_active}"
            self.column_range_label.setText(f"{row_info}    |    {col_info}")
        else: self.column_range_label.setText("   Rows: 0\n   Columns: 0")


    @Slot(pl.DataFrame)
    def on_filter_completed(self, filtered_df_result):
        self.filtered_df = filtered_df_result # This should already have __original_index__
        
        self.df_shape_after = filtered_df_result.shape if filtered_df_result is not None else None

        logger.log_action(
            "Pickaxe", "Structured Filter Applied", 
            f"Applied {len(self.applied_filters_info)} filter conditions.", # self.applied_filters_info should store filter tuples/logic
            details=self.applied_filters_info, # This will be a list of tuples/strings
            df_shape_before=self.df_shape_before, # Shape of df *before* this filter pass
            df_shape_after=self.df_shape_after # Shape of df *after* this filter pass
        )
        
        self.df_shape_before = None
        self.df_shape_after = None

        self.update_model_slot(self.filtered_df, True)
        if self.table_view.selectionModel().hasSelection():
            current_selection = self.table_view.selectionModel().currentIndex()
            if current_selection.isValid(): self.handle_column_selection_for_stats(current_selection, None)
        else: self.stats_label.setText("Select a column to see statistics."); self.filled_rows_bar.set_fill_percentage(0); self.set_stats_label_font_monospaced()
        self.update_progress(100, "Filtering completed.")

    def toggle_filter_panel(self):
        if self.filter_panel:
            self.filter_panel.setVisible(not self.filter_panel.isVisible())
            if self.filter_panel.isVisible():
                self.update_query_input_from_structured_filters()


    def set_row_as_header(self):
        if self.df is None: return
        selected_indexes = self.table_view.selectionModel().selectedIndexes()
        if not selected_indexes :
            QMessageBox.information(self, "Set Header", "Please select a cell in the row to be set as header."); return

        # view_row_index is the visual row index in the QTableView.
        # We need the corresponding __original_index__ from that visual row.
        view_row_index = selected_indexes[0].row()
        if self.model and hasattr(self.model, '_original_indices') and view_row_index < len(self.model._original_indices):
            target_original_df_row_index = self.model._original_indices[view_row_index]
        else:
            QMessageBox.warning(self, "Set Header Error", "Could not determine original row index for header.")
            return

        if target_original_df_row_index >= self.df.height:
            QMessageBox.warning(self, "Set Header Error", "Selected row index is out of bounds for the base dataset.")
            self.update_progress(100,"Error setting header")
            return

        self.update_progress(0, "Setting new header...")
        QApplication.processEvents()
        try:
            # Get header values from the *original* self.df using target_original_df_row_index
            # Exclude __original_index__ when fetching header values
            cols_for_header_vals = [c for c in self.df.columns if c != "__original_index__"]
            new_header_series = self.df.select(cols_for_header_vals).row(target_original_df_row_index)
            new_header_values_from_df = [str(x) for x in new_header_series]

            # The new DataFrame should also exclude __original_index__ initially, then re-add it.
            df_for_reheader = self.df.drop("__original_index__")
            df_before_reheader_shape = self.df.shape 
            
            if len(new_header_values_from_df) != len(df_for_reheader.columns):
                QMessageBox.warning(self, "Set Header Error", f"Header length mismatch. Got {len(new_header_values_from_df)} new headers for {len(df_for_reheader.columns)} columns.")
                self.update_progress(100,"Error setting header")
                return

            rename_mapping = {old_col: new_val for old_col, new_val in zip(df_for_reheader.columns, new_header_values_from_df)}

            self.df = df_for_reheader.slice(target_original_df_row_index + 1).rename(rename_mapping)
            self.df = self.df.with_row_count("__original_index__") # Re-add original index to the new df
            self.original_row_count = self.df.height # Update based on new df
            self.filtered_df = self.df.clone()
            self.applied_filters_info = []; self.query_input.clear()

            self.update_model_slot(self.filtered_df, True)
            self.update_completer_model() # Headers changed

            if self.filter_panel:
                self.filter_panel.panel_filters_changed.disconnect(self.update_query_input_from_structured_filters)
                self.filter_panel.deleteLater()

            new_panel_cols = [c for c in self.df.columns if c != "__original_index__"]
            self.filter_panel = FilterPanel(new_panel_cols)
            self.filter_panel.panel_filters_changed.connect(self.update_query_input_from_structured_filters)
            self.filter_panel.apply_button.clicked.connect(self.apply_structured_filters)
            self._layout.insertWidget(4, self.filter_panel)
            self.filter_panel.setVisible(False); self.filter_panel.setEnabled(True); self.filter_toggle_button.setEnabled(True)
            self.update_query_input_from_structured_filters()

            self.update_progress(100, "New header set successfully.")
            
            logger.log_action(
                "Pickaxe", "Header Changed", ...,
                df_shape_before=df_before_reheader_shape, df_shape_after=self.df.shape
            )

        except Exception as e:
            self.update_progress(100, f"Error setting header: {e}")
            QMessageBox.critical(self, "Set Header Error", str(e))
            logger.log_action( "Pickaxe", "Header Changing Error", "Error Setting Header", details={"Error": str(e)})

    def update_file_info_label(self):
        if not self._filepath or not self.file_info: self.file_info_label.setText("No file loaded."); return
        file_name_base = os.path.basename(self._filepath)
        display_file_name = file_name_base if len(file_name_base) <= 60 else file_name_base[:25] + "..." + file_name_base[-30:]
        file_info_text = f"  File: '{display_file_name}'\n"
        if self._filepath.name.lower().endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
            file_info_text += f" Sheet: '{self.file_info.get('sheet_name', 'N/A')}' (Total: {self.file_info.get('total_sheets', 'N/A')}) | "
        else: file_info_text += f" Delimiter: '{self.file_info.get('delimiter', 'N/A')}' | Encoding: '{self.file_info.get('encoding', 'N/A')}' | "

        # file_info['rows'] and ['columns'] are from original load before __original_index__
        original_rows = self.file_info.get('rows', 'N/A')
        original_cols = self.file_info.get('columns', 'N/A')
        file_info_text += f"Original Loaded: {original_rows:,} R x {original_cols:,} C | "


        file_size_mb = self.file_info.get('size', 0)
        if file_size_mb > 0:
            if file_size_mb > 1024: file_info_text += f"Size: {file_size_mb/1024:,.2f} GB | "
            elif file_size_mb >= 1: file_info_text += f"Size: {file_size_mb:,.2f} MB | "
            else: file_info_text += f"Size: {file_size_mb*1024:,.2f} KB | "
        load_time_sec = self.file_info.get('load_time', 0)
        if load_time_sec > 0:
            if load_time_sec > 60: mins, secs = int(load_time_sec // 60), load_time_sec % 60; file_info_text += f"Load time: {mins}m {secs:.2f}s"
            else: file_info_text += f"Load time: {load_time_sec:.2f}s"
        else: file_info_text += f"Load time: N/A"
        self.file_info_label.setText(file_info_text.strip())

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            focused_widget = QApplication.focusWidget()
            
            if focused_widget and self.filter_panel:
                foc_widg_flg = self.filter_panel.isAncestorOf(focused_widget)
            else:
                foc_widg_flg = False
                
            if self.filter_panel and self.filter_panel.isVisible() and foc_widg_flg:
                if isinstance(focused_widget, (QLineEdit, QComboBox, QCheckBox, QRadioButton, QDateEdit)):
                    self.apply_structured_filters(); event.accept(); return
            elif focused_widget == self.query_input:
                self.apply_polars_query(); event.accept(); return
        super().keyPressEvent(event)

    def get_current_dataframe_for_vw(self):
        """Returns the current DataFrame (filtered or original) for Visual Workshop."""
        if self.filtered_df is not None and not self.filtered_df.is_empty():
            return self.filtered_df.clone()
        elif self.df is not None and not self.df.is_empty():
            return self.df.clone()
        return None

    def get_current_filename_hint_for_vw(self):
        return self._filepath if self._filepath else "Pickaxe Data"

    def open_visual_workshop(self):
        self.statusBar().showMessage("Opening Visual Workshop...", 5000)

        if self.visual_workshop_window is None or not self.visual_workshop_window.isVisible():
            from visual_workshop import VisualWorkshopApp # Local import
            self.visual_workshop_window = VisualWorkshopApp(pickaxe_instance=self, log_file_to_use=self.current_log_file_path)
        
        if not self.types_suggested_and_applied_this_session and self.df is not None:
            reply = QMessageBox.question(self, "Type Suggestions",
                                       "Data types have not been reviewed in this session.\n"
                                       "Would you like to run type suggestions before opening Visual Workshop?\n"
                                       "Should you like to run type suggestions, click 'Yes' and open Visual Workshop afterwards.\n"
                                       "(This can improve plotting accuracy for some columns).",
                                       QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                                       QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                self.suggest_and_convert_types()
            elif reply == QMessageBox.StandardButton.Cancel:
                self.statusBar().showMessage("Visual Workshop launch cancelled.", 2000)
                return
        else:
            self.continue_opening_visual_workshop()

    def continue_opening_visual_workshop(self):
        """Continues the process of opening Visual Workshop after type suggestions."""
        current_df = self.get_current_dataframe_for_vw()
        if current_df is None:
            QMessageBox.information(self, "No Data", "No data loaded in Pickaxe to send to Visual Workshop.")
            return

        file_hint = self.get_current_filename_hint_for_vw()
        if self.current_log_file_path: logger.set_log_file(self.current_log_file_path, "Pickaxe")
        logger.log_action("Pickaxe", "Launch External Tool", "Visual Workshop launched.", 
                            details={"Data Hint": file_hint, "Data Shape": current_df.shape if current_df is not None else "N/A"})

        # Always send the current data and log path when showing
        if self.visual_workshop_window:
            self.visual_workshop_window.receive_dataframe(current_df, file_hint, log_file_path_from_source=self.current_log_file_path)
            self.visual_workshop_window.show()
            self.visual_workshop_window.raise_()
            self.visual_workshop_window.activateWindow()

    def closeEvent(self, event):
        # First, handle potential child windows like VW and DT
        if self.visual_workshop_window and self.visual_workshop_window.isVisible():
            # Close VW, which should have its own "save?" mechanism if needed
            self.visual_workshop_window.close() 
            if self.visual_workshop_window.isVisible(): # If VW cancel its close
                event.ignore()
                return

        if self.data_transformer_window and self.data_transformer_window.isVisible():
            # Close DT, which should also have its own "save?" mechanism
            self.data_transformer_window.close()
            if self.data_transformer_window.isVisible(): # If DT cancel its close
                event.ignore()
                return

        # Now, check if Pickaxe itself has data that might need saving
        # A more robust check would be if 'self.df' has unsaved changes, 
        # but for now, we'll prompt if data is loaded.
        if self.df is not None and not self.df.is_empty(): # Check if there's data
            reply = QMessageBox.question(self, 'Confirm Exit',
                                           "Do you want to save your current data before closing Pickaxe?",
                                           QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No | QMessageBox.StandardButton.Cancel,
                                           QMessageBox.StandardButton.Cancel) # Default to Cancel

            if reply == QMessageBox.StandardButton.Yes:
                # Attempt to save. self.save_file() opens a dialog.
                # We need to know if the save was successful or if the user cancelled the save dialog.
                # Modify save_file to return a status or check if _filepath was updated.
                # For simplicity here, we assume save_file will handle its own flow.
                # If save_file itself is cancelled by user, it won't proceed.
                # A more robust way is for save_file to return True on success, False on cancel.
                
                # Let's assume save_file shows dialog and saves. If user cancels save dialog,
                # they effectively don't want to save.
                # We need a way for save_file to indicate if it proceeded or was cancelled by user.
                # For now, we'll call save_file. If it completes without error, we assume it's okay to close.
                # This is a simplification. A better save_file would return a status.
                
                # Simple approach: if save_file is called, and they *don't* cancel its dialog, it saves.
                # This prompt is about *whether to initiate* the save process.
                
                # Let's make save_file return a boolean
                save_success = self.save_file_for_close_event() # A new method to handle this
                if save_success is None: # User cancelled the "do you want to save" dialog itself
                    event.ignore() # This case is handled by QMessageBox.StandardButton.Cancel
                elif save_success: # Save was successful OR user chose not to save in the save_file dialog
                    logger.log_action("Pickaxe", "Application Close", "User chose to save (or cancelled save dialog) and then closed.")
                    event.accept()
                else: # Save was attempted but failed, or save_file itself indicated a reason not to close (e.g. explicit cancel)
                    event.ignore() # Don't close if save failed or was explicitly cancelled from save_file dialog

            elif reply == QMessageBox.StandardButton.No:
                logger.log_action("Pickaxe", "Application Close", "User chose not to save and closed.")
                event.accept() # Close without saving
            else: # QMessageBox.StandardButton.Cancel or user closed the dialog
                logger.log_action("Pickaxe", "Application Close", "User cancelled closing.")
                event.ignore() # Don't close
        else:
            logger.log_action("Pickaxe", "Application Close", "Application closed (no data to save).")
            event.accept() # No data to save, close normally
        # super().closeEvent(event) # No, event.accept() or event.ignore() is sufficient.
        
    def save_file_for_close_event(self):
        """
        Dedicated save function for closeEvent to handle dialog cancellation.
        Returns:
            True if saved successfully or if user cancelled the save *dialog* (implying they are ok to proceed with closing without save).
            False if an error occurred during save, or if a more explicit "don't close" is needed from save.
            For simplicity, if the QFileDialog is cancelled, we treat it as "user doesn't want to save this way".
        """
        if self.df is None:
            return True # Nothing to save

        original_name = os.path.splitext(os.path.basename(self._filepath if self._filepath else "data"))[0]
        # ... (filter_suffix and suggested_name logic from your save_file)
        filter_info_str_parts = []
        for item in self.applied_filters_info:
            if isinstance(item, tuple):
                column, comparison, fs1, fs2, _, ftype, regex, negate = item
                fs1_clean = re.sub(r'[^a-zA-Z0-9_-]', '', str(fs1))[:10]
                fs2_clean = re.sub(r'[^a-zA-Z0-9_-]', '', str(fs2))[:10] if fs2 else ""
                comp_short = comparison.replace(" ", "")[:10].lower() 
                col_str = str(column)[:6] if column else "Row"
                part = f"{col_str}_{'N' if negate else ''}{'R' if regex else ''}{ftype[:1]}_{comp_short}_{fs1_clean}"
                if fs2_clean: part += f"_{fs2_clean}"
                filter_info_str_parts.append(part)
            elif isinstance(item, str) and item.startswith("Query: "):
                query_clean = re.sub(r'[^a-zA-Z0-9_=-]', '', item.replace("Query: ", ""))[:20]
                filter_info_str_parts.append(f"Q_{query_clean}")
            elif isinstance(item, str): filter_info_str_parts.append(item)
        filter_suffix = "_".join(filter_info_str_parts)
        filter_suffix = re.sub(r'_+', '_', filter_suffix).strip('_')
        suggested_name = f"{original_name}_processed.csv" if not filter_suffix else f"{original_name}_filtered_{filter_suffix}.csv"
        max_len_suggested = 200
        if len(suggested_name) > max_len_suggested:
            ext = os.path.splitext(suggested_name)[1]
            suggested_name = suggested_name[:max_len_suggested - len(ext)] + ext

        file_name, selected_filter = QFileDialog.getSaveFileName(self, "Save Current Data", suggested_name, 
                                                                 "CSV Files (*.csv);;Excel Files (*.xlsx)")
        if file_name:
            self.progress_bar.setVisible(True)
            self.progress_label.setVisible(True)
            self.progress_bar.setValue(0)
            
            df_to_save_op = self.filtered_df if self.filtered_df is not None and not self.filtered_df.is_empty() else self.df
            if df_to_save_op is None:
                QMessageBox.warning(self, "Save Error", "No data available to save.")
                self.progress_bar.setVisible(False); self.progress_label.setVisible(False)
                return False # Indicate save failed / aborted

            # Log before starting thread, assuming save will proceed
            if self.current_log_file_path:
                 logger.set_log_file(self.current_log_file_path, "Pickaxe")
            logger.log_action("Pickaxe", "File Save Initiated", f"Attempting to save data to '{os.path.basename(file_name)}'.",
                details={"Path": file_name, "Format": selected_filter, 
                         "Source Data": os.path.basename(self._filepath) if self._filepath else "Current Data"})


            # Using FileSaverWorker implies async save. For closeEvent, a synchronous save might be better
            # or the closeEvent logic needs to handle the async nature (which is complex).
            # For simplicity here, let's assume a direct synchronous save for the close event's save prompt.
            try:
                temp_df_to_save = df_to_save_op.clone()
                if "__original_index__" in temp_df_to_save.columns:
                    temp_df_to_save = temp_df_to_save.drop("__original_index__")

                if file_name.lower().endswith('.csv'):
                    dialect = self.file_info.get('dialect', {'delimiter': ',', 'quotechar': '"'})
                    sep = dialect.get('delimiter', ',')
                    quote_char = dialect.get('quotechar', '"')
                    temp_df_to_save.write_csv(file_name, separator=sep, quote_char=quote_char, quote_style="non_numeric")
                elif file_name.lower().endswith('.xlsx'):
                    temp_df_to_save.write_excel(file_name, autofit=True)
                
                self.statusBar().showMessage(f"File saved successfully as {file_name}", 3000)
                logger.log_dataframe_save(
                    "Pickaxe", file_name, 
                    rows=temp_df_to_save.height, cols=temp_df_to_save.width,
                    source_data_name=os.path.basename(self._filepath) if self._filepath else "current data"
                )
                return True # Save successful
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Error saving file: {str(e)}")
                logger.log_action("Pickaxe", "File Save Error", f"Error saving to '{os.path.basename(file_name)}'.", details={"Error": str(e)})
                return False # Save failed
            finally:
                self.progress_bar.setVisible(False); self.progress_label.setVisible(False)
        else:
            return True # User cancelled the QFileDialog, implies they don't want to save, so allow closing.


    def open_data_transformer(self):
        self.statusBar().showMessage("Opening Data Transformer...", 2000)
        current_df = self._get_df_to_operate_on() # Get current (filtered or original) df
        if current_df is None:
            QMessageBox.information(self, "No Data", "Load data in Pickaxe first.")
            return
        
        file_hint = self.get_current_filename_hint_for_vw()
        if self.current_log_file_path: logger.set_log_file(self.current_log_file_path, "Pickaxe")
        logger.log_action("Pickaxe", "Launch External Tool", "Data Transformer launched.",
                             details={"Data Hint": file_hint, "Data Shape": current_df.shape if current_df is not None else "N/A"})

        if self.data_transformer_window is None or not self.data_transformer_window.isVisible():
            # self.statusBar().showMessage("Launching Data Transformer...", 2000)
            from data_transformer import DataTransformer
            self.data_transformer_window = DataTransformer(
                source_app=self, 
                initial_df=current_df.clone(), 
                df_name_hint=file_hint, 
                log_file_to_use=self.current_log_file_path # Pass the path
            )
        else:
            self.statusBar().showMessage("Data Transformer already open. Sending data...", 2000)
            self.data_transformer_window.receive_dataframe(current_df.clone(), file_hint, log_file_path_from_source=self.current_log_file_path)

        
        self.data_transformer_window.show()
        self.data_transformer_window.raise_()
        self.data_transformer_window.activateWindow()
        # If DT already had data, ask if user wants to replace it or open new instance (optional)
        # For simplicity, current behavior is to pass data to existing instance if open
        if hasattr(self.data_transformer_window, 'receive_dataframe'): # Check if method exists
            self.data_transformer_window.receive_dataframe(current_df.clone(), self._filepath or "Pickaxe Data")
